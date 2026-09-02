"""
Parser voor FPBB_doelenboom_referentietabel_*.xlsx workbooks.

Ondersteunt twee formaten, automatisch herkend aan de aanwezige tabbladen:
- "oud": de huidige productiestructuur, 9 tabbladen zoals gedocumenteerd in
  doelenboom_update_instructie.md §2 (Referentietabel, Capability-OB relaties,
  Project-Capability relaties, Projecten, Producten, Tags, Element-Tag relaties,
  Organisatieonderdelen, OB-Organisatie relaties).
- "nieuw": het voorstel uit voorstel_excel_structuur_v2.md — geen aparte
  Capability-OB/Project-Capability-tabbladen en geen "Bovenliggend element"-kolom,
  maar één generieke "Relaties"-tab (Bron-ID, Doel-ID, Relatietype, Toelichting)
  met alle relaties (verticaal én horizontaal). Referentietabel heeft geen
  statuskolommen meer maar wel "Volgorde" en "Actief". Herkenning: workbook heeft
  een tabblad "Relaties" i.p.v. "Capability-OB relaties"/"Project-Capability relaties".

Beide formaten worden door dezelfde Referentietabel-/relatielogica hieronder
verwerkt: de kolomherkenning is alias-gebaseerd, dus kolommen die in het ene
formaat niet bestaan (bv. "Bovenliggend element" in het nieuwe formaat, of
"Volgorde"/"Actief" in het oude) komen gewoon leeg/None terug en worden genegeerd.

Kolomnamen worden alias-gebaseerd herkend (case/whitespace-ongevoelig) i.p.v. op
exacte string, omdat kleine headernaam-verschillen tussen Excel-versies eerder al
zijn voorgekomen. Elke rij die niet eenduidig te interpreteren is (onbekend type,
ontbrekende FK-referentie, onherkend statuslabel) wordt overgeslagen en als warning
in het rapport gezet — nooit geraden, conform de "eerst melden"-regel uit de
instructie. De aanroeper (API) toont dit rapport aan de gebruiker voordat er
gepubliceerd kan worden (zie routes/imports.ts: upload -> report -> publish).

Publiceren (routes/imports.ts, POST /api/imports/:id/publish) is en blijft een
volledige vervanging: alle elements/edges/tags/org_units van de doelenboom worden
eerst verwijderd en daarna opnieuw ingevoegd vanuit de geparste data. Een rij die
in de nieuwe upload ontbreekt — of in het nieuwe formaat op Actief = "Nee" staat —
komt dus niet meer terug en verdwijnt bij publiceren definitief uit de database.
"""
from __future__ import annotations

import io
import re
from collections import OrderedDict, defaultdict
from typing import Any

from openpyxl import load_workbook

from .cleaning import clean_date, clean_pct, clean_text, norm, split_entries
from .dependency_format import parse_dependency_entry

# --- Type-normalisatie (Referentietabel "Type"-kolom -> canonieke DB-waarde) ---
# De canonieke set komt uit de check-constraint in db/init.sql.
TYPE_MAP: dict[str, str] = {
    'project': 'Project',
    'capability': 'Capability',
    'operationele benefit': 'Operationele benefit',
    'operationele baat': 'Operationele benefit',  # variant zoals in doelenboom.html-brondata
    'sub-benefit': 'Sub-benefit',
    'sub-benefit fpbb': 'Sub-benefit',
    'subbenefit': 'Sub-benefit',
    'programmabaat': 'Programmabaat',
    'programmabaat fpbb': 'Programmabaat',
    'strategische benefit': 'Strategische benefit',
    'strategisch benefit': 'Strategische benefit',
    'strategisch benefit b&b': 'Strategische benefit',
    'strategisch doel': 'Strategisch doel',
    'missie': 'Missie',
    'missie kmar': 'Missie',
}
# Types die altijd genegeerd worden (traceability-archief, geen actuele portfolio —
# zie doelenboom_update_instructie.md §2a).
SKIP_TYPES = {'project (historisch)'}

# Types waarvoor een "verticale" edge automatisch afgeleid mag worden uit de
# "Bovenliggend element"-kolom, als die tekst exact een bekende elementcode is.
# Project->Capability en Capability->Operationele benefit komen altijd uit de
# relatietabbladen, niet uit deze kolom.
VERTICAL_PARENT_TYPES = {
    'Operationele benefit', 'Sub-benefit', 'Programmabaat',
    'Strategische benefit', 'Strategisch doel',
}

RELATIETYPE_MAP = {
    'primair': 'primair',
    'ondersteunend': 'ondersteunend',
}

PROJECTSTATUS_VALUES = {'backlog', 'actief', 'on-hold', 'gereed', 'vervallen'}
RAG_MAP = {'rood': 'Rood', 'oranje': 'Oranje', 'groen': 'Groen'}
# "Planning item"-type van een product/deliverable — zie db/init.sql (products.type).
PRODUCT_TYPE_MAP = {'deliverable': 'deliverable', 'mijlpaal': 'mijlpaal', 'milestone': 'mijlpaal'}


def find_col(headers: list[Any], *aliases: str) -> int | None:
    normalized = [norm(h) for h in headers]
    for alias in aliases:
        a = norm(alias)
        for i, h in enumerate(normalized):
            if h == a:
                return i
    return None


def read_sheet(wb, sheet_name: str, field_aliases: dict[str, tuple[str, ...]]):
    """Geeft (rows, found) terug. rows is een lijst van dicts met de logische
    veldnamen uit field_aliases als key. found is False als het tabblad ontbreekt."""
    if sheet_name not in wb.sheetnames:
        return [], False
    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(row_iter))
    except StopIteration:
        return [], True

    col_index: dict[str, int | None] = {
        field: find_col(headers, *aliases) for field, aliases in field_aliases.items()
    }

    rows = []
    for raw_row in row_iter:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        row = {}
        for field, idx in col_index.items():
            row[field] = raw_row[idx] if idx is not None and idx < len(raw_row) else None
        rows.append(row)
    return rows, True


def read_config_value(wb, sheet_name: str, key: str) -> str | None:
    """Leest één waarde uit een 'Sleutel'/'Waarde'-tabblad zoals de
    Configuratie-tab (exporter.py) of Info-tab (project_workbook.py) — o.a.
    gebruikt om de export-formaatversie terug te lezen (zie
    excel_format_version.py). Geeft None als het tabblad of de sleutel
    ontbreekt, i.p.v. te falen — oudere bestanden van vóór de versie-kolom
    moeten gewoon blijven werken, alleen dan zonder versie in het rapport."""
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for raw_row in ws.iter_rows(values_only=True):
        if not raw_row or len(raw_row) < 2:
            continue
        cell_key = raw_row[0]
        if cell_key is not None and norm(str(cell_key)) == norm(key):
            value = raw_row[1]
            return str(value).strip() if value not in (None, '') else None
    return None


def parse_workbook(
    content: bytes, filename: str = '', valid_types: list[str] | None = None
) -> tuple[str, dict, dict | None]:
    warnings: list[str] = []
    errors: list[str] = []
    sheets_found: list[str] = []
    sheets_missing: list[str] = []

    # Welke Type-waarden geldig zijn hangt sinds de configureerbare kolommen
    # (zie docs/kolommen-configuratie-ontwerp.md) af van de kolomconfiguratie
    # van de doelenboom waarin geïmporteerd wordt — de aanroeper (routes/imports.ts)
    # geeft die hier altijd mee. Zonder valid_types (bv. een oudere aanroep, of
    # los getest) vallen we terug op de historische, vaste set uit TYPE_MAP (de
    # 8 standaardtypes) — zelfde gedrag als vóór de configureerbare kolommen.
    allowed_types = set(valid_types) if valid_types is not None else set(TYPE_MAP.values())
    allowed_by_norm = {norm(t): t for t in allowed_types}

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - we rapporteren dit terug, geen crash
        return 'failed', {
            'errors': [f'Kon workbook niet openen: {exc}'],
            'warnings': [], 'counts': {}, 'sheetsFound': [], 'sheetsMissing': [],
        }, None

    def track(sheet_name: str, found: bool):
        (sheets_found if found else sheets_missing).append(sheet_name)

    # Formaatherkenning: het nieuwe formaat heeft één "Relaties"-tab i.p.v. de
    # aparte "Capability-OB relaties"/"Project-Capability relaties"-tabbladen.
    detected_format = 'nieuw' if 'Relaties' in wb.sheetnames else 'oud'

    # --- Referentietabel (verplicht) ---
    # Bevat de velden van beide formaten tegelijk; een kolom die in het gebruikte
    # formaat niet bestaat (bv. "Bovenliggend element" in het nieuwe formaat, of
    # "Volgorde"/"Actief" in het oude) levert gewoon None op per rij.
    ref_rows, ref_found = read_sheet(wb, 'Referentietabel', {
        'id': ('id', 'element-id', 'code'),
        'type': ('type',),
        'naam': ('element', 'naam', 'elementnaam'),
        'beschrijving': ('uitgebreide beschrijving', 'beschrijving'),
        'bovenliggend': ('bovenliggend element', 'parent'),
        'kpi': ('kpi', 'mogelijke kpi / indicator', 'mogelijke kpi/indicator'),
        'taakveld': ('taakveld',),
        'subtaakveld': ('sub-taakveld', 'subtaakveld'),
        'projectstatus': ('projectstatus', 'status'),
        'rag': ('rag-status', 'rag'),
        'statustoelichting': ('statustoelichting',),
        'gerapporteerd_op': ('status gerapporteerd op', 'gerapporteerd op'),
        'volgorde': ('volgorde',),
        'actief': ('actief',),
    })
    track('Referentietabel', ref_found)
    if not ref_found:
        errors.append('Tabblad "Referentietabel" ontbreekt — dit is de enige verplichte bron voor elementen.')
        return 'failed', {
            'errors': errors, 'warnings': warnings, 'counts': {},
            'sheetsFound': sheets_found, 'sheetsMissing': sheets_missing,
        }, None

    elements: list[dict] = []
    code_set: set[str] = set()
    parent_by_code: dict[str, str] = {}
    type_by_code: dict[str, str] = {}
    ref_status: dict[str, dict] = {}
    skipped_historisch = 0
    skipped_inactief = 0
    normalized_type_labels: dict[str, int] = defaultdict(int)
    sort_counter = 0

    for row in ref_rows:
        code = clean_text(row.get('id'))
        naam = clean_text(row.get('naam'))
        if not code or not naam:
            continue
        # Nieuw formaat: een element met Actief = "Nee" wordt behandeld als
        # verwijderd — het wordt niet meegenomen in de geparste data, en verdwijnt
        # dus bij publiceren (volledige vervanging) definitief uit de database,
        # net als een rij die helemaal uit de Referentietabel is verwijderd.
        actief_raw = clean_text(row.get('actief'))
        if actief_raw and norm(actief_raw) in ('nee', 'n', 'no', 'false'):
            skipped_inactief += 1
            continue
        type_raw = clean_text(row.get('type'))
        type_key = norm(type_raw)
        if type_key in SKIP_TYPES:
            skipped_historisch += 1
            continue
        type_norm = TYPE_MAP.get(type_key)
        if type_norm is not None and type_norm not in allowed_types:
            # Een TYPE_MAP-alias (bv. "operationele baat" -> "Operationele
            # benefit") die bij déze doelenboom geen geldige kolom (meer) is
            # (de tenant heeft de kolommen aangepast) — niet alsnog accepteren.
            type_norm = None
        if type_norm is None:
            # Geen (geldige) TYPE_MAP-alias: probeer een directe, case/
            # whitespace-ongevoelige match tegen de geconfigureerde type-namen
            # van deze doelenboom — nodig voor eigen, niet-standaard kolommen.
            type_norm = allowed_by_norm.get(type_key)
        if type_norm is None:
            warnings.append(f'Onbekend Type-label "{type_raw}" bij element {code} — rij overgeslagen.')
            continue
        if type_norm != type_raw:
            normalized_type_labels[f'{type_raw} -> {type_norm}'] += 1
        if code in code_set:
            warnings.append(f'Dubbele elementcode "{code}" in Referentietabel — alleen de eerste rij is gebruikt.')
            continue

        code_set.add(code)
        type_by_code[code] = type_norm
        sort_counter += 1
        # Nieuw formaat heeft een expliciete Volgorde-kolom; die krijgt voorrang
        # boven de leesvolgorde-teller. Oud formaat heeft geen Volgorde-kolom, dus
        # blijft altijd bij de teller.
        sort_order = sort_counter
        volgorde_raw = row.get('volgorde')
        if volgorde_raw not in (None, ''):
            try:
                sort_order = int(volgorde_raw)
            except (TypeError, ValueError):
                warnings.append(
                    f'Ongeldige Volgorde-waarde "{volgorde_raw}" bij element {code} — leesvolgorde gebruikt.'
                )
        parent_text = clean_text(row.get('bovenliggend'))
        if parent_text:
            parent_by_code[code] = parent_text
        elements.append({
            'code': code,
            'type': type_norm,
            'name': naam,
            'description': clean_text(row.get('beschrijving')),
            'parentText': parent_text,
            'kpi': clean_text(row.get('kpi')),
            'taakveld': clean_text(row.get('taakveld')),
            'subtaakveld': clean_text(row.get('subtaakveld')),
            'sortOrder': sort_order,
        })

        # Referentietabel kan (sinds v15, voor Project-rijen) ook statuskolommen
        # bevatten — die krijgen voorrang boven de Projecten-tab (zie §2a).
        ps_val = clean_text(row.get('projectstatus'))
        rag_val = clean_text(row.get('rag'))
        toel_val = clean_text(row.get('statustoelichting'))
        datum_val = clean_date(row.get('gerapporteerd_op'))
        if ps_val or rag_val or toel_val or datum_val:
            ref_status[code] = {
                'projectstatus': ps_val, 'rag': rag_val,
                'toelichting': toel_val, 'gerapporteerdOp': datum_val,
            }

    if skipped_historisch:
        warnings.append(
            f'{skipped_historisch} rij(en) met Type "Project (historisch)" genegeerd '
            '(traceability-archief, geen actuele portfolio — zie §2a).'
        )
    if skipped_inactief:
        warnings.append(
            f'{skipped_inactief} element(en) met Actief = "Nee" genegeerd — deze worden bij '
            'publiceren verwijderd als ze eerder al in deze doelenboom bestonden.'
        )
    for label, count in normalized_type_labels.items():
        warnings.append(f'Type-label "{label}" genormaliseerd voor {count} rij(en).')

    # --- Relatietabbladen ---
    # cap_ob/proj_cap horen alleen bij het oude formaat en Relaties alleen bij het
    # nieuwe; alleen de tabbladen die bij het gedetecteerde formaat horen worden
    # meegeteld in sheetsFound/sheetsMissing, anders zou een geldige nieuw-formaat
    # upload ten onrechte "ontbrekende tabbladen" tonen (en andersom).
    cap_ob_rows, cap_ob_found = read_sheet(wb, 'Capability-OB relaties', {
        'source': ('capability-id', 'capability id'),
        'target': ('ob-id', 'ob id'),
        'relatietype': ('relatietype',),
        'toelichting': ('toelichting',),
    })
    proj_cap_rows, proj_cap_found = read_sheet(wb, 'Project-Capability relaties', {
        'source': ('project-id', 'project id'),
        'target': ('capability-id', 'capability id'),
        'relatietype': ('relatietype',),
        'toelichting': ('toelichting',),
    })
    if detected_format == 'oud':
        track('Capability-OB relaties', cap_ob_found)
        track('Project-Capability relaties', proj_cap_found)

    edges: list[dict] = []
    seen_edges: set[tuple[str, str]] = set()
    missing_fk = 0

    def add_relation_edges(rows: list[dict], label: str):
        nonlocal missing_fk
        for row in rows:
            source = clean_text(row.get('source'))
            target = clean_text(row.get('target'))
            if not source or not target:
                continue
            if source not in code_set or target not in code_set:
                missing_fk += 1
                warnings.append(
                    f'{label}: relatie {source} -> {target} verwijst naar een onbekende elementcode — overgeslagen.'
                )
                continue
            rel_key = norm(row.get('relatietype'))
            weight = RELATIETYPE_MAP.get(rel_key)
            if weight is None and rel_key:
                warnings.append(f'{label}: onbekend Relatietype "{row.get("relatietype")}" bij {source} -> {target}.')
            if (source, target) in seen_edges:
                continue
            seen_edges.add((source, target))
            edges.append({
                'source': source, 'target': target, 'weight': weight,
                'toelichting': clean_text(row.get('toelichting')),
            })

    add_relation_edges(cap_ob_rows, 'Capability-OB relaties')
    add_relation_edges(proj_cap_rows, 'Project-Capability relaties')

    # Nieuw formaat: één generieke "Relaties"-tab i.p.v. de twee tabbladen hierboven,
    # met zowel de vroegere Capability-OB/Project-Capability-relaties als de
    # verticale relaties die in het oude formaat uit "Bovenliggend element" werden
    # afgeleid (zie de _fill_nieuw-export in exporter.py). Dezelfde validatie
    # (onbekende elementcode, onbekend Relatietype) geldt hiervoor.
    rel_rows, rel_found = read_sheet(wb, 'Relaties', {
        'source': ('bron-id', 'bron id'),
        'target': ('doel-id', 'doel id'),
        'relatietype': ('relatietype',),
        'toelichting': ('toelichting',),
    })
    if detected_format == 'nieuw':
        track('Relaties', rel_found)
    if rel_found:
        add_relation_edges(rel_rows, 'Relaties')

    # Verticale edges (OB -> Sub -> Programmabaat -> Strategische benefit -> Strategisch doel -> Missie),
    # afgeleid uit "Bovenliggend element". Voor de meeste niveaus staat daar gewoon een
    # exacte elementcode. Vanaf Programmabaat opwaarts blijkt de kolom in de praktijk
    # ook samengestelde tekst te bevatten, bv. "A1 primair; A2 ondersteunend" of
    # "A1/A3 ondersteunend" (meerdere ouders, optioneel met Relatietype-woord erachter),
    # en voor Strategisch doel simpelweg het woord "Missie" i.p.v. de code M1 (er is er
    # toch maar één). Dit wordt hieronder structureel geparsed i.p.v. genegeerd — alles
    # wat na het parsen nog niet naar een bekende code herleid kan worden, wordt gemeld.
    type_name_to_single_code: dict[str, str] = {}
    type_counts: dict[str, int] = defaultdict(int)
    for c, t in type_by_code.items():
        type_counts[t] += 1
    for c, t in type_by_code.items():
        if type_counts[t] == 1:
            type_name_to_single_code[norm(t)] = c

    def resolve_parent_code(text: str) -> str | None:
        if text in code_set:
            return text
        return type_name_to_single_code.get(norm(text))

    weight_suffix_re = re.compile(r'^(?P<codes>.+?)\s+(?P<weight>primair|ondersteunend)$', re.IGNORECASE)

    vertical_unresolved = 0
    unresolved_examples: list[str] = []
    for code, parent_text in parent_by_code.items():
        if type_by_code.get(code) not in VERTICAL_PARENT_TYPES:
            continue
        for segment in parent_text.split(';'):
            segment = segment.strip()
            if not segment:
                continue
            m = weight_suffix_re.match(segment)
            if m:
                codes_part, weight = m.group('codes').strip(), m.group('weight').lower()
            else:
                codes_part, weight = segment, None
            for raw_code in codes_part.split('/'):
                raw_code = raw_code.strip()
                if not raw_code:
                    continue
                resolved = resolve_parent_code(raw_code)
                if resolved is None:
                    vertical_unresolved += 1
                    if len(unresolved_examples) < 8:
                        unresolved_examples.append(f'{code}: "{raw_code}" (uit "{parent_text}")')
                    continue
                key = (resolved, code)
                if key not in seen_edges:
                    seen_edges.add(key)
                    edges.append({'source': resolved, 'target': code, 'weight': weight, 'toelichting': ''})
    if vertical_unresolved:
        warnings.append(
            f'{vertical_unresolved} verwijzing(en) in "Bovenliggend element" konden niet naar een bekende '
            'elementcode herleid worden, ook niet na het splitsen op ";"/"/" en Relatietype-woord: '
            + '; '.join(unresolved_examples) + ('; ...' if vertical_unresolved > len(unresolved_examples) else '')
        )

    # --- Projecten (status/naam, fallback t.o.v. Referentietabel) ---
    proj_rows, proj_found = read_sheet(wb, 'Projecten', {
        'id': ('project-id', 'id'),
        'projectstatus': ('projectstatus', 'status'),
        'rag': ('rag-status', 'rag'),
        'statustoelichting': ('statustoelichting',),
        'gerapporteerd_op': ('status gerapporteerd op', 'gerapporteerd op'),
        'cluster_ppt': ('cluster ppt', 'cluster'),
        'beschrijving': ('uitgebreide beschrijving', 'beschrijving'),
    })
    track('Projecten', proj_found)

    project_status: dict[str, dict] = dict(ref_status)  # Referentietabel heeft voorrang
    description_fallback: dict[str, str] = {}
    for row in proj_rows:
        code = clean_text(row.get('id'))
        if not code or code not in code_set:
            continue
        fallback = {
            'projectstatus': clean_text(row.get('projectstatus')),
            'rag': clean_text(row.get('rag')),
            'toelichting': clean_text(row.get('statustoelichting')),
            'gerapporteerdOp': clean_date(row.get('gerapporteerd_op')),
            'clusterPpt': clean_text(row.get('cluster_ppt')),
        }
        existing = project_status.get(code, {})
        merged = {k: (existing.get(k) or fallback.get(k)) for k in ('projectstatus', 'rag', 'toelichting', 'gerapporteerdOp', 'clusterPpt')}
        project_status[code] = merged

        proj_desc = clean_text(row.get('beschrijving'))
        if proj_desc:
            description_fallback[code] = proj_desc

    # "Uitgebreide beschrijving" op de Projecten-tab is alleen een fallback: als de
    # Referentietabel-rij van dit project al een beschrijving had, blijft die leidend.
    if description_fallback:
        applied = 0
        for el in elements:
            if el['type'] == 'Project' and not el['description'] and el['code'] in description_fallback:
                el['description'] = description_fallback[el['code']]
                applied += 1
        if applied:
            warnings.append(
                f'{applied} project(en) zonder beschrijving in Referentietabel: "Uitgebreide beschrijving" '
                'van de Projecten-tab gebruikt als fallback.'
            )

    # Zorg dat elk project een clusterPpt-veld heeft, ook als er geen Projecten-rij was.
    for code in project_status:
        project_status[code].setdefault('clusterPpt', '')

    # Normaliseer/valideer statuswaarden tegen de DB check-constraints.
    for code, ps in list(project_status.items()):
        ps_key = norm(ps.get('projectstatus'))
        if ps_key and ps_key not in PROJECTSTATUS_VALUES:
            warnings.append(f'Onbekende Projectstatus "{ps["projectstatus"]}" bij {code} — leeggelaten.')
            ps['projectstatus'] = ''
        elif ps_key:
            ps['projectstatus'] = ps_key.capitalize()  # 'on-hold' -> 'On-hold', 'actief' -> 'Actief', ...
        rag_key = norm(ps.get('rag'))
        if rag_key:
            if rag_key in RAG_MAP:
                ps['rag'] = RAG_MAP[rag_key]
            else:
                warnings.append(f'Onbekende RAG-status "{ps["rag"]}" bij {code} — leeggelaten.')
                ps['rag'] = ''

    # --- Producten ---
    prod_rows, prod_found = read_sheet(wb, 'Producten', {
        'id': ('product-id', 'id'),
        'project_id': ('project-id',),
        'naam': ('product / deliverable', 'product/deliverable', 'product', 'deliverable', 'naam'),
        'type': ('type',),
        'omschrijving': ('omschrijving',),
        # 'voortgang (0-100)' erbij: de kolomkop van het "nieuw" Excel-formaat
        # (zie exporter.py NIEUW_SHEET_HEADERS) — zonder deze alias kwam
        # pctGereed bij elke import van dat formaat altijd als 0 terug, omdat
        # geen van de andere aliassen ('% gereed' e.d., alleen gebruikt in het
        # "oud" formaat) matchte. Gevonden via de geautomatiseerde round-trip-
        # regressietest (tests/test_roundtrip.py).
        'pct': ('% gereed', 'pct gereed', 'gereed', 'voortgang (0-100)', 'voortgang'),
        'verwacht': ('verwachte opleverdatum', 'verwachte datum'),
        'werkelijk': ('werkelijke opleverdatum', 'werkelijke datum'),
        'opmerking': ('opmerking',),
    })
    track('Producten', prod_found)

    products: dict[str, list[dict]] = defaultdict(list)
    for row in prod_rows:
        project_code = clean_text(row.get('project_id'))
        naam = clean_text(row.get('naam'))
        if not project_code or not naam:
            continue
        if project_code not in code_set:
            warnings.append(f'Producten: product "{naam}" verwijst naar onbekend Project-ID "{project_code}" — overgeslagen.')
            continue
        type_raw = clean_text(row.get('type'))
        type_key = norm(type_raw)
        if type_key and type_key in PRODUCT_TYPE_MAP:
            product_type = PRODUCT_TYPE_MAP[type_key]
        else:
            if type_key:
                warnings.append(f'Producten: onbekend Type "{type_raw}" bij "{naam}" — gebruikt "deliverable".')
            product_type = 'deliverable'
        products[project_code].append({
            'code': clean_text(row.get('id')),
            'name': naam,
            'type': product_type,
            'omschrijving': clean_text(row.get('omschrijving')),
            'pctGereed': clean_pct(row.get('pct')),
            'verwachteDatum': clean_date(row.get('verwacht')),
            'werkelijkeDatum': clean_date(row.get('werkelijk')),
            'opmerking': clean_text(row.get('opmerking')),
        })

    # --- Activiteiten ---
    # Ontbreekt de tab (een export van vóór deze tab bestond), dan blijft
    # 'activities' gewoon leeg — net als bij Producten hierboven is dit geen
    # verplicht tabblad (zie REQUIRED_SHEETS-achtige check, die alleen op
    # Referentietabel let), dus geen enkele oudere upload breekt hierop.
    # Voorgangers verwijzen naar activiteitnamen BINNEN hetzelfde project (zie
    # dependency_format.parse_dependency_entry) — net als bij Producten/
    # elementen wordt de eigenlijke db-id pas bij publiceren toegekend
    # (routes/imports.ts: volledige vervanging), dus namen zijn hier de enige
    # zinvolle sleutel; de aanroeper lost 'name' op naar de nieuwe id ná
    # het aanmaken van alle activiteiten van dat project.
    act_rows, act_found = read_sheet(wb, 'Activiteiten', {
        'project_id': ('project-id',),
        'naam': ('activiteit', 'naam'),
        'start': ('startdatum',),
        'eind': ('einddatum',),
        'omschrijving': ('omschrijving',),
        'mijlpaal': ('mijlpaal',),
        'fase': ('fase/samenvattend', 'fase', 'samenvattend'),
        'voorgangers': ('voorgangers',),
    })
    track('Activiteiten', act_found)

    activities: dict[str, list[dict]] = defaultdict(list)
    for row in act_rows:
        project_code = clean_text(row.get('project_id'))
        naam = clean_text(row.get('naam'))
        start = clean_date(row.get('start'))
        eind = clean_date(row.get('eind'))
        if not project_code or not naam:
            continue
        if project_code not in code_set:
            warnings.append(f'Activiteiten: activiteit "{naam}" verwijst naar onbekend Project-ID "{project_code}" — overgeslagen.')
            continue
        if not start or not eind:
            warnings.append(f'Activiteiten: activiteit "{naam}" mist een start- en/of einddatum — overgeslagen.')
            continue
        predecessors = []
        for entry in split_entries(row.get('voorgangers')):
            pred_name, dep_type, lag_days = parse_dependency_entry(entry)
            if pred_name:
                predecessors.append({'name': pred_name, 'type': dep_type, 'lagDays': lag_days})
        activities[project_code].append({
            'name': naam,
            'startDate': start,
            'endDate': eind,
            'omschrijving': clean_text(row.get('omschrijving')),
            'isMilestone': norm(clean_text(row.get('mijlpaal'))) == 'ja',
            'isSummary': norm(clean_text(row.get('fase'))) == 'ja',
            'predecessors': predecessors,
        })

    # --- Tags + Element-Tag relaties ---
    tag_rows, tags_found = read_sheet(wb, 'Tags', {
        'id': ('tag-id', 'id'),
        'naam': ('tag', 'naam'),
        'categorie': ('categorie',),
        'omschrijving': ('omschrijving',),
    })
    track('Tags', tags_found)

    tags: list[dict] = []
    tag_code_set: set[str] = set()
    for row in tag_rows:
        code = clean_text(row.get('id'))
        naam = clean_text(row.get('naam'))
        if not code or not naam or code in tag_code_set:
            continue
        tag_code_set.add(code)
        tags.append({
            'code': code, 'name': naam,
            'categorie': clean_text(row.get('categorie')),
            'omschrijving': clean_text(row.get('omschrijving')),
        })

    et_rows, et_found = read_sheet(wb, 'Element-Tag relaties', {
        'element_id': ('element-id',),
        'tag_id': ('tag-id',),
        'toelichting': ('toelichting',),
    })
    track('Element-Tag relaties', et_found)

    element_tags: dict[str, list[str]] = defaultdict(list)
    for row in et_rows:
        element_code = clean_text(row.get('element_id'))
        tag_code = clean_text(row.get('tag_id'))
        if not element_code or not tag_code:
            continue
        if element_code not in code_set:
            warnings.append(f'Element-Tag relaties: onbekende elementcode "{element_code}" — overgeslagen.')
            continue
        if tag_code not in tag_code_set:
            warnings.append(f'Element-Tag relaties: onbekende Tag-ID "{tag_code}" — overgeslagen.')
            continue
        if tag_code not in element_tags[element_code]:
            element_tags[element_code].append(tag_code)

    # --- Organisatieonderdelen + OB-Organisatie relaties ---
    org_rows, org_found = read_sheet(wb, 'Organisatieonderdelen', {
        'id': ('org-id', 'id'),
        'naam': ('organisatieonderdeel', 'naam'),
        'omschrijving': ('omschrijving',),
    })
    track('Organisatieonderdelen', org_found)

    org_units: list[dict] = []
    org_code_set: set[str] = set()
    for row in org_rows:
        code = clean_text(row.get('id'))
        naam = clean_text(row.get('naam'))
        if not code or not naam or code in org_code_set:
            continue
        org_code_set.add(code)
        org_units.append({'code': code, 'name': naam, 'omschrijving': clean_text(row.get('omschrijving'))})

    obo_rows, obo_found = read_sheet(wb, 'OB-Organisatie relaties', {
        'ob_id': ('ob-id',),
        'org_id': ('org-id',),
        'relatietype': ('relatietype',),
        'toelichting': ('toelichting',),
        'status': ('status',),
    })
    track('OB-Organisatie relaties', obo_found)

    ob_org: dict[str, list[dict]] = defaultdict(list)
    for row in obo_rows:
        ob_code = clean_text(row.get('ob_id'))
        org_code = clean_text(row.get('org_id'))
        if not ob_code or not org_code:
            continue  # hulpkolommen-rij (dropdown-brondata), geen echte relatie — zie §2
        if ob_code not in code_set:
            warnings.append(f'OB-Organisatie relaties: onbekende OB-code "{ob_code}" — overgeslagen.')
            continue
        if org_code not in org_code_set:
            warnings.append(f'OB-Organisatie relaties: onbekende Org-ID "{org_code}" — overgeslagen.')
            continue
        status = clean_text(row.get('status')) or 'Concept'
        if status not in ('Concept', 'Gevalideerd', 'Vervallen'):
            warnings.append(f'OB-Organisatie relaties: onbekende Status "{status}" bij {ob_code}/{org_code} — "Concept" gebruikt.')
            status = 'Concept'
        ob_org[ob_code].append({
            'org': org_code,
            'relatietype': clean_text(row.get('relatietype')) or 'Betrokken',
            'toelichting': clean_text(row.get('toelichting')),
            'status': status,
        })

    counts = {
        'elements': len(elements),
        'edges': len(edges),
        'projectStatus': len(project_status),
        'products': sum(len(v) for v in products.values()),
        'activities': sum(len(v) for v in activities.values()),
        'tags': len(tags),
        'elementTags': sum(len(v) for v in element_tags.values()),
        'orgUnits': len(org_units),
        'obOrg': sum(len(v) for v in ob_org.values()),
        'skippedHistorisch': skipped_historisch,
        'skippedInactief': skipped_inactief,
        'missingForeignKeys': missing_fk,
    }

    parsed = {
        'elements': elements,
        'edges': edges,
        'projectStatus': project_status,
        'products': dict(products),
        'activities': dict(activities),
        'tags': tags,
        'elementTags': dict(element_tags),
        'orgUnits': org_units,
        'obOrg': dict(ob_org),
    }

    status = 'ok'
    if warnings:
        status = 'warning'
    if not elements:
        errors.append('Referentietabel bevat geen bruikbare elementrijen.')
        status = 'failed'

    report = {
        'filename': filename,
        'format': detected_format,
        'formatVersion': read_config_value(wb, 'Configuratie', 'Export-formaatversie'),
        'sheetsFound': sheets_found,
        'sheetsMissing': sheets_missing,
        'counts': counts,
        'warnings': warnings,
        'errors': errors,
    }
    return status, report, (parsed if status != 'failed' else None)
