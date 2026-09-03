"""
Export/import van de volledige gegevens van ÉÉN project (Project-element) als
Excel-bestand — drie tabbladen (Project, Producten, Activiteiten), plus een
puur informatieve Info-tab. Dit is bewust een ander, veel kleiner formaat dan
de hele-doelenboom export/import in exporter.py/parser.py hierboven: dit gaat
over precies één project (bv. "Sweepen"), niet over de hele boom.

Round-trip-ontwerp: elk product/elke activiteit krijgt een 'ID'-kolom met de
databank-id (bigserial uit products/activities). Bij opnieuw importeren wordt
op die ID gematcht — dat gebeurt hier NIET, dat is aan de aanroeper (zie
computeProjectImportPlan in tree.html): een lege ID betekent "nieuwe rij", een
bestaande ID betekent "bijwerken", en een product/activiteit van dit project
waarvan de ID nergens in het bestand meer voorkomt is een kandidaat om te
verwijderen. Dat is betrouwbaarder dan naam-matching, en eenvoudiger dan de
mppUid-aanpak bij de MS Project-import (dat bronbestand komt nooit met ónze
eigen database-ID's terug — dit Excel-bestand, als het een re-import van een
eerdere export is, wél).

Deze module doet alleen de rauwe Excel<->JSON-vertaling: het bepalen van
create/update/delete (het "wijzigingsoverzicht" dat de gebruiker moet
bevestigen, zie de projectExcelRouter-toelichting in
api/src/routes/projectExcel.ts) gebeurt client-side, net als bij de MS
Project-import.
"""
from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .cleaning import clean_date, clean_pct, clean_text, norm, split_entries as _split_entries
from .dependency_format import format_dependency as _format_dependency
from .dependency_format import format_product_dependency as _format_product_dependency
from .dependency_format import parse_dependency_entry as _parse_dependency_entry
from .dependency_format import parse_product_dependency_entry as _parse_product_dependency_entry
from .excel_format_version import PROJECT_EXPORT_FORMAT_VERSION
from .exporter import VALIDATIELIJSTEN
from .parser import read_config_value, read_sheet
from .xlsx_safety import create_safe_sheet

INFO_SHEET = 'Info'
PROJECT_SHEET = 'Project'
PRODUCTS_SHEET = 'Producten'
ACTIVITIES_SHEET = 'Activiteiten'

PROJECT_HEADERS = [
    'Code', 'Naam', 'Omschrijving', 'Projectstatus', 'RAG-status',
    'Statustoelichting', 'Status gerapporteerd op', 'Cluster PPT',
    'Tags', 'Organisatieonderdelen',
]
PRODUCT_HEADERS = [
    'ID', 'Naam', 'Type', 'Omschrijving', '% gereed',
    'Verwachte opleverdatum', 'Werkelijke opleverdatum', 'Deadline',
    'Duur', 'Eenheid', 'BV', 'Opmerking', 'Hangt af van',
]
ACTIVITY_HEADERS = [
    'ID', 'Naam', 'Startdatum', 'Einddatum', 'Omschrijving',
    'Mijlpaal', 'Fase/samenvattend', 'Voorgangers',
]

REQUIRED_SHEETS = (PROJECT_SHEET, PRODUCTS_SHEET, ACTIVITIES_SHEET)

# ---------------------------------------------------------------------------
# Bouwen (export)
# ---------------------------------------------------------------------------


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'


def _set_widths(ws: Worksheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _add_dropdown(ws: Worksheet, values: list[str], cell_range: str) -> None:
    # Letterlijke lijst in de formule i.p.v. een verwijzing naar een aparte
    # validatielijsten-tab (zoals exporter.py voor Type doet) — hier gaat het
    # steeds om een handvol korte, vaste waarden, dus is een los tabblad
    # onnodige overhead. Excel's limiet voor zo'n inline lijst is 255 tekens;
    # alle lijsten hieronder (Deliverable/Mijlpaal, Ja/Nee, d/w/m/y,
    # Projectstatus, RAG-status) blijven daar ruim onder.
    formula = '"' + ','.join(values) + '"'
    dv = DataValidation(type='list', formula1=formula, allow_blank=True, showErrorMessage=True)
    dv.error = 'Kies een waarde uit de lijst.'
    dv.errorTitle = 'Ongeldige waarde'
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_project_workbook(data: dict[str, Any], meta: dict[str, Any]) -> bytes:
    project = data.get('project') or {}
    products = data.get('products') or []
    product_deps = data.get('productDependencies') or []
    activities = data.get('activities') or []
    activity_deps = data.get('activityDependencies') or []

    product_name_by_id = {p.get('id'): p.get('name', '') for p in products}
    activity_name_by_id = {a.get('id'): a.get('name', '') for a in activities}

    product_preds: dict[Any, list[tuple[str, int, str]]] = {}
    for d in product_deps:
        product_preds.setdefault(d.get('successorId'), []).append((
            product_name_by_id.get(d.get('predecessorId'), '?'),
            d.get('lagAmount') or 0,
            d.get('lagEenheid') or 'd',
        ))
    activity_preds: dict[Any, list[tuple[str, str, int]]] = {}
    for d in activity_deps:
        activity_preds.setdefault(d.get('successorId'), []).append(
            (activity_name_by_id.get(d.get('predecessorId'), '?'), d.get('type') or 'FS', d.get('lagDays') or 0)
        )

    wb = Workbook()
    wb.remove(wb.active)

    info_ws = create_safe_sheet(wb, INFO_SHEET)
    info_ws.append(['Sleutel', 'Waarde'])
    for cell in info_ws[1]:
        cell.font = Font(bold=True)
    for key, value in [
        ('Project', project.get('name', '')),
        ('Projectcode', project.get('code', '')),
        ('Doelenboom', meta.get('doelenboom', '')),
        ('Tenant', meta.get('tenant', '')),
        ('Export-formaatversie', PROJECT_EXPORT_FORMAT_VERSION),
        ('Geëxporteerd op', meta.get('exportedAt', '')),
        ('Geëxporteerd door', meta.get('exportedBy', '')),
    ]:
        info_ws.append([key, value])
    _set_widths(info_ws, [22, 44])

    proj_ws = create_safe_sheet(wb, PROJECT_SHEET)
    _write_header(proj_ws, PROJECT_HEADERS)
    status = project.get('status') or {}
    tags_str = '; '.join(project.get('tags') or [])
    orgs_str = '; '.join(
        f"{o.get('name', '')} ({o.get('relatietype', '')})" for o in (project.get('orgs') or [])
    )
    proj_ws.append([
        project.get('code', ''), project.get('name', ''), project.get('description', ''),
        status.get('projectstatus') or '', status.get('rag') or '',
        status.get('toelichting') or '', status.get('gerapporteerdOp') or '',
        status.get('clusterPpt') or '', tags_str, orgs_str,
    ])
    _add_dropdown(proj_ws, VALIDATIELIJSTEN['Projectstatus'], 'D2:D2')
    _add_dropdown(proj_ws, VALIDATIELIJSTEN['RAG-status'], 'E2:E2')
    _set_widths(proj_ws, [12, 26, 40, 14, 12, 30, 18, 16, 24, 34])

    prod_ws = create_safe_sheet(wb, PRODUCTS_SHEET)
    _write_header(prod_ws, PRODUCT_HEADERS)
    for p in products:
        prod_ws.append([
            p.get('id'), p.get('name', ''),
            'Mijlpaal' if p.get('type') == 'mijlpaal' else 'Deliverable',
            p.get('omschrijving', ''), p.get('pctGereed') or 0,
            p.get('verwachteDatum') or '', p.get('werkelijkeDatum') or '',
            p.get('deadline') or '',
            p.get('duur') if p.get('duur') is not None else '',
            p.get('duurEenheid') or 'd',
            p.get('businessValue') if p.get('businessValue') is not None else '',
            p.get('opmerking', ''),
            '; '.join(_format_product_dependency(n, a, e) for (n, a, e) in product_preds.get(p.get('id'), [])),
        ])
    _add_dropdown(prod_ws, ['Deliverable', 'Mijlpaal'], 'C2:C10000')
    _add_dropdown(prod_ws, ['d', 'w', 'm', 'y'], 'J2:J10000')
    _set_widths(prod_ws, [8, 30, 12, 34, 10, 16, 16, 14, 8, 9, 8, 30, 34])

    act_ws = create_safe_sheet(wb, ACTIVITIES_SHEET)
    _write_header(act_ws, ACTIVITY_HEADERS)
    for a in activities:
        preds = activity_preds.get(a.get('id'), [])
        act_ws.append([
            a.get('id'), a.get('name', ''), a.get('startDate') or '', a.get('endDate') or '',
            a.get('omschrijving', ''),
            'Ja' if a.get('isMilestone') else 'Nee',
            'Ja' if a.get('isSummary') else 'Nee',
            '; '.join(_format_dependency(n, t, l) for (n, t, l) in preds),
        ])
    _add_dropdown(act_ws, ['Ja', 'Nee'], 'F2:F10000')
    _add_dropdown(act_ws, ['Ja', 'Nee'], 'G2:G10000')
    _set_widths(act_ws, [8, 30, 14, 14, 34, 10, 16, 40])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------------------------------------------------------------------------
# Parsen (import) — geeft (status, report, parsed) terug, zelfde contract als
# parse_workbook in parser.py.
# ---------------------------------------------------------------------------

def _parse_id(value: object) -> int | None:
    if value is None or value == '':
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def parse_project_workbook(content: bytes) -> tuple[str, dict, dict | None]:
    warnings: list[str] = []
    errors: list[str] = []
    sheets_found: list[str] = []
    sheets_missing: list[str] = []

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — teruggeven i.p.v. laten crashen
        return 'failed', {
            'errors': [f'Kon workbook niet openen: {exc}'],
            'warnings': [], 'counts': {}, 'sheetsFound': [], 'sheetsMissing': [],
        }, None

    for name in REQUIRED_SHEETS:
        (sheets_found if name in wb.sheetnames else sheets_missing).append(name)
    if sheets_missing:
        errors.append(
            'Verplichte tabblad(en) ontbreken: ' + ', '.join(sheets_missing) +
            ' — is dit een export uit Doelenboom (Bestand > Project exporteren als Excel)?'
        )
        return 'failed', {
            'errors': errors, 'warnings': warnings, 'counts': {},
            'sheetsFound': sheets_found, 'sheetsMissing': sheets_missing,
        }, None

    project_rows, _ = read_sheet(wb, PROJECT_SHEET, {
        'code': ('code',), 'naam': ('naam',), 'omschrijving': ('omschrijving',),
        'projectstatus': ('projectstatus',), 'rag': ('rag-status', 'rag'),
        'toelichting': ('statustoelichting',), 'gerapporteerd_op': ('status gerapporteerd op',),
        'cluster_ppt': ('cluster ppt',), 'tags': ('tags',),
        'orgs': ('organisatieonderdelen',),
    })
    project: dict[str, Any] = {}
    if project_rows:
        row = project_rows[0]
        rag_raw = clean_text(row.get('rag'))
        projectstatus_raw = clean_text(row.get('projectstatus'))
        # Zelfde validatie als de rest van de app (project-status.ts /
        # PROJECTSTATUS_VALUES-parser hierboven) — een onbekende waarde wordt
        # gemeld en leeggelaten i.p.v. geraden of geweigerd; de rest van het
        # bestand blijft gewoon bruikbaar.
        if projectstatus_raw and norm(projectstatus_raw) not in {'backlog', 'actief', 'on-hold', 'gereed', 'vervallen'}:
            warnings.append(f'Project: onbekende Projectstatus "{projectstatus_raw}" — leeggelaten.')
            projectstatus_raw = ''
        if rag_raw and norm(rag_raw) not in {'rood', 'oranje', 'groen'}:
            warnings.append(f'Project: onbekende RAG-status "{rag_raw}" — leeggelaten.')
            rag_raw = ''
        orgs: list[dict[str, str]] = []
        for entry in _split_entries(row.get('orgs')):
            m = re.match(r'^(?P<name>.*?)\s*\((?P<rol>[^()]+)\)\s*$', entry)
            if m:
                orgs.append({'name': m.group('name').strip(), 'relatietype': m.group('rol').strip()})
            else:
                orgs.append({'name': entry, 'relatietype': ''})
        project = {
            'code': clean_text(row.get('code')),
            'name': clean_text(row.get('naam')),
            'description': clean_text(row.get('omschrijving')),
            'projectstatus': projectstatus_raw.lower() if projectstatus_raw else '',
            'rag': rag_raw.lower() if rag_raw else '',
            'toelichting': clean_text(row.get('toelichting')),
            'gerapporteerdOp': clean_date(row.get('gerapporteerd_op')),
            'clusterPpt': clean_text(row.get('cluster_ppt')),
            'tags': _split_entries(row.get('tags')),
            'orgs': orgs,
        }
    else:
        warnings.append('Tabblad "Project" bevat geen datarij — projectgegevens worden overgeslagen.')

    product_rows, _ = read_sheet(wb, PRODUCTS_SHEET, {
        'id': ('id',), 'naam': ('naam',), 'type': ('type',), 'omschrijving': ('omschrijving',),
        'pct': ('% gereed', 'pct gereed'), 'verwacht': ('verwachte opleverdatum',),
        'werkelijk': ('werkelijke opleverdatum',), 'deadline': ('deadline',),
        'duur': ('duur',), 'eenheid': ('eenheid',), 'bv': ('bv', 'business value'),
        'opmerking': ('opmerking',), 'hangt_af_van': ('hangt af van',),
    })
    products: list[dict[str, Any]] = []
    for row in product_rows:
        naam = clean_text(row.get('naam'))
        if not naam:
            continue
        type_raw = clean_text(row.get('type'))
        product_type = 'mijlpaal' if norm(type_raw) == 'mijlpaal' else 'deliverable'
        eenheid_raw = norm(clean_text(row.get('eenheid')))
        eenheid = eenheid_raw if eenheid_raw in ('d', 'w', 'm', 'y') else 'd'
        if eenheid_raw and eenheid_raw not in ('d', 'w', 'm', 'y'):
            warnings.append(f'Producten: onbekende Eenheid "{row.get("eenheid")}" bij "{naam}" — "d" gebruikt.')
        duur_raw = row.get('duur')
        duur = None
        if duur_raw not in (None, ''):
            try:
                duur = int(float(duur_raw))
            except (TypeError, ValueError):
                warnings.append(f'Producten: ongeldige Duur "{duur_raw}" bij "{naam}" — leeggelaten.')
        bv_raw = row.get('bv')
        business_value = None
        if bv_raw not in (None, ''):
            try:
                business_value = float(bv_raw)
            except (TypeError, ValueError):
                warnings.append(f'Producten: ongeldige BV "{bv_raw}" bij "{naam}" — leeggelaten.')
        depends_on = []
        for entry in _split_entries(row.get('hangt_af_van')):
            dep_name, lag_amount, lag_eenheid = _parse_product_dependency_entry(entry)
            if dep_name:
                depends_on.append({'name': dep_name, 'lagAmount': lag_amount, 'lagEenheid': lag_eenheid})
        products.append({
            'id': _parse_id(row.get('id')),
            'name': naam,
            'type': product_type,
            'omschrijving': clean_text(row.get('omschrijving')),
            'pctGereed': clean_pct(row.get('pct')),
            'verwachteDatum': clean_date(row.get('verwacht')),
            'werkelijkeDatum': clean_date(row.get('werkelijk')),
            'deadline': clean_date(row.get('deadline')),
            'duur': duur,
            'duurEenheid': eenheid,
            'businessValue': business_value,
            'opmerking': clean_text(row.get('opmerking')),
            'dependsOn': depends_on,
        })

    activity_rows, _ = read_sheet(wb, ACTIVITIES_SHEET, {
        'id': ('id',), 'naam': ('naam',), 'start': ('startdatum',), 'eind': ('einddatum',),
        'omschrijving': ('omschrijving',), 'mijlpaal': ('mijlpaal',),
        'fase': ('fase/samenvattend', 'fase', 'samenvattend'), 'voorgangers': ('voorgangers',),
    })
    activities: list[dict[str, Any]] = []
    for row in activity_rows:
        naam = clean_text(row.get('naam'))
        start = clean_date(row.get('start'))
        eind = clean_date(row.get('eind'))
        if not naam or not start or not eind:
            warnings.append(f'Activiteiten: rij overgeslagen (naam/start-/einddatum ontbreekt) — "{naam or "?"}".')
            continue
        predecessors = []
        for entry in _split_entries(row.get('voorgangers')):
            name, dep_type, lag_days = _parse_dependency_entry(entry)
            if name:
                predecessors.append({'name': name, 'type': dep_type, 'lagDays': lag_days})
        activities.append({
            'id': _parse_id(row.get('id')),
            'name': naam,
            'startDate': start,
            'endDate': eind,
            'omschrijving': clean_text(row.get('omschrijving')),
            'isMilestone': norm(clean_text(row.get('mijlpaal'))) == 'ja',
            'isSummary': norm(clean_text(row.get('fase'))) == 'ja',
            'predecessors': predecessors,
        })

    counts = {'products': len(products), 'activities': len(activities)}
    status = 'ok'
    if warnings:
        status = 'warning'
    if not project and not products and not activities:
        errors.append('Geen bruikbare gegevens gevonden in dit bestand.')
        status = 'failed'

    report = {
        'errors': errors, 'warnings': warnings, 'counts': counts,
        'sheetsFound': sheets_found, 'sheetsMissing': sheets_missing,
        'formatVersion': read_config_value(wb, INFO_SHEET, 'Export-formaatversie'),
    }
    parsed = {'project': project, 'products': products, 'activities': activities}
    return status, report, (parsed if status != 'failed' else None)
