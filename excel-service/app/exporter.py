"""
Bouwt een FPBB_doelenboom_referentietabel-achtige .xlsx terug vanuit de
TreeResponse-vorm die de API teruggeeft bij GET /api/doelenbomen/:id/tree.

Twee formaten:
- "oud": de huidige productiestructuur (9 tabbladen, zoals
  FPBB_doelenboom_referentietabel_v15.xlsx). Kolomkoppen zijn 1-op-1 overgenomen
  van een echte upload zodat dit bestand zonder aanpassing weer via /parse
  ingelezen kan worden (round-trip).
- "nieuw": het voorstel uit voorstel_excel_structuur_v2.md — Capability-OB
  relaties, Project-Capability relaties én de vrije-tekst "Bovenliggend
  element"-relaties vanaf Programmabaat worden vervangen door één generieke
  Relaties-tab (Bron-ID, Doel-ID, Relatietype, Toelichting); Referentietabel
  bevat geen statuskolommen meer (die staan alleen nog op Projecten) en krijgt
  een expliciete Volgorde- en Actief-kolom; plus een _Validatielijsten-tab met
  dropdown-brondata en bijbehorende Data Validation op alle gesloten-lijstvelden.
  parser.py detecteert dit formaat automatisch (aanwezigheid van de
  Relaties-tab) en kan het net als "oud" weer inlezen.

Beide formaten krijgen een Configuratie-tab (doelenboom, tenant, formaat, modus,
export-formaatversie, geëxporteerd op/door) — nuttig om te kunnen zien waar een
bestand vandaan komt, ook als het los rondgestuurd wordt — en een Kolommen-tab
met de volledige
kolomconfiguratie (volgorde, type, titel, ondertitel, kleur, smal, projectrol,
relatielabel, lettergrootte) van déze doelenboom op het moment van exporteren.

Twee modi (onafhankelijk van het formaat):
- "template": alleen tabbladen + headers, geen datarijen.
- "data": headers + één rij per record uit de huidige databron.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from .dependency_format import format_dependency, format_product_dependency
from .excel_format_version import TREE_EXPORT_FORMAT_VERSION
from .xlsx_safety import create_safe_sheet

CONFIG_SHEET_NAME = 'Configuratie'
VALIDATIELIJSTEN_SHEET_NAME = '_Validatielijsten'
KOLOMMEN_SHEET_NAME = 'Kolommen'

# ---------------------------------------------------------------------------
# Oud formaat — huidige productiestructuur
# ---------------------------------------------------------------------------
OUD_SHEET_HEADERS: dict[str, list[str]] = {
    'Referentietabel': [
        'ID', 'Type', 'Element', 'Uitgebreide beschrijving', 'Bovenliggend element',
        'Mogelijke KPI / indicator', 'Taakveld', 'Sub-taakveld',
        'Projectstatus', 'RAG-status', 'Statustoelichting', 'Status gerapporteerd op',
    ],
    'Capability-OB relaties': ['Capability-ID', 'Capability', 'OB-ID', 'Operationele benefit', 'Relatietype', 'Toelichting'],
    'Project-Capability relaties': ['Project-ID', 'Project', 'Capability-ID', 'Capability', 'Relatietype', 'Toelichting'],
    'Projecten': [
        'Project-ID', 'Project', 'Cluster PPT', 'Projectstatus', 'RAG-status',
        'Statustoelichting', 'Status gerapporteerd op', 'Uitgebreide beschrijving',
    ],
    'Producten': [
        'Product-ID', 'Project-ID', 'Project', 'Product / deliverable', 'Type', 'Omschrijving',
        '% gereed', 'Verwachte opleverdatum', 'Werkelijke opleverdatum', 'Opmerking', 'Hangt af van',
    ],
    'Activiteiten': [
        'Activiteit-ID', 'Project-ID', 'Project', 'Activiteit', 'Startdatum', 'Einddatum',
        'Omschrijving', 'Mijlpaal', 'Fase/samenvattend', 'Voorgangers',
    ],
    'Tags': ['Tag-ID', 'Tag', 'Categorie', 'Omschrijving'],
    'Element-Tag relaties': ['Element-ID', 'Type', 'Element', 'Tag-ID', 'Tag', 'Toelichting'],
    'Organisatieonderdelen': ['Org-ID', 'Organisatieonderdeel', 'Omschrijving'],
    'OB-Organisatie relaties': ['OB-ID', 'Operationele benefit', 'Org-ID', 'Organisatieonderdeel', 'Relatietype', 'Toelichting', 'Status'],
}
OUD_SHEET_ORDER = list(OUD_SHEET_HEADERS.keys())

# ---------------------------------------------------------------------------
# Nieuw formaat — voorstel_excel_structuur_v2.md
# ---------------------------------------------------------------------------
NIEUW_SHEET_HEADERS: dict[str, list[str]] = {
    'Referentietabel': [
        'ID', 'Type', 'Element', 'Uitgebreide beschrijving',
        'Mogelijke KPI / indicator', 'Taakveld', 'Sub-taakveld', 'Volgorde', 'Actief',
    ],
    'Relaties': ['Bron-ID', 'Doel-ID', 'Relatietype', 'Toelichting'],
    'Projecten': [
        'Project-ID', 'Project', 'Cluster PPT', 'Projectstatus', 'RAG-status',
        'Statustoelichting', 'Status gerapporteerd op', 'Uitgebreide beschrijving',
    ],
    'Producten': [
        'Product-ID', 'Project-ID', 'Project', 'Product / deliverable', 'Type', 'Omschrijving',
        'Voortgang (0-100)', 'Verwachte opleverdatum', 'Werkelijke opleverdatum', 'Opmerking', 'Hangt af van',
    ],
    'Activiteiten': [
        'Activiteit-ID', 'Project-ID', 'Project', 'Activiteit', 'Startdatum', 'Einddatum',
        'Omschrijving', 'Mijlpaal', 'Fase/samenvattend', 'Voorgangers',
    ],
    'Tags': ['Tag-ID', 'Tag', 'Categorie', 'Omschrijving'],
    'Element-Tag relaties': ['Element-ID', 'Type', 'Element', 'Tag-ID', 'Tag', 'Toelichting'],
    'Organisatieonderdelen': ['Org-ID', 'Organisatieonderdeel', 'Omschrijving'],
    'OB-Organisatie relaties': ['OB-ID', 'Operationele benefit', 'Org-ID', 'Organisatieonderdeel', 'Relatietype', 'Toelichting', 'Status'],
}
NIEUW_SHEET_ORDER = list(NIEUW_SHEET_HEADERS.keys())

# De 8 standaardtypes, in hun oorspronkelijke volgorde — zie ook
# standardColumns() in api/src/columnConfig.ts (bewust twee keer uitgeschreven,
# zelfde reden als daar: TS-runtime-code en deze Python-service raken elkaar
# niet, geen gedeeld pad zonder dat een van beide de ander als afhankelijkheid
# zou moeten inladen). Gebruikt om (a) het "oud" formaat te beperken tot
# doelenbomen die nog exact deze standaardconfig hebben (zie is_standard_columns
# hieronder, en docs/kolommen-configuratie-ontwerp.md) en (b) als fallback-
# Type-lijst wanneer geen columns zijn meegegeven (bv. een oudere aanroep).
STANDARD_TYPE_NAMES: list[str] = [
    'Project', 'Capability', 'Operationele benefit', 'Sub-benefit',
    'Programmabaat', 'Strategische benefit', 'Strategisch doel', 'Missie',
]


def is_standard_columns(columns: list[dict[str, Any]]) -> bool:
    """True als de meegegeven kolommen (op position gesorteerd) qua type-namen
    en volgorde exact overeenkomen met de 8 standaardkolommen — titel/ondertitel/
    kleur mogen wél afwijken (bv. een tenantnaam in de titel), dat verandert
    niets aan de structuur waar het "oud" Excel-formaat van uitgaat."""
    ordered = sorted(columns, key=lambda c: c.get('position', 0))
    return [c.get('typeName') for c in ordered] == STANDARD_TYPE_NAMES


def _type_names_from_columns(columns: list[dict[str, Any]]) -> list[str]:
    """Type-namen van de meegegeven kolommen, op position gesorteerd — de
    dynamische vervanger van de vroeger hardgecodeerde VALIDATIELIJSTEN['Type'].
    Valt terug op STANDARD_TYPE_NAMES als er geen columns zijn meegegeven
    (defensief; de aanroeper (main.py) geeft deze normaliter altijd mee)."""
    if not columns:
        return list(STANDARD_TYPE_NAMES)
    ordered = sorted(columns, key=lambda c: c.get('position', 0))
    names = [c.get('typeName') for c in ordered if c.get('typeName')]
    return names or list(STANDARD_TYPE_NAMES)


# Dropdown-brondata voor het nieuwe formaat (§4 en §7 van het voorstel). 'Type'
# staat hier bewust NIET meer vast in — die wordt per export dynamisch bepaald
# uit de kolomconfiguratie van de doelenboom (zie _type_names_from_columns en
# _write_validatielijsten hieronder), want dat is nu per tenant/doelenboom
# configureerbaar (zie docs/kolommen-configuratie-ontwerp.md).
VALIDATIELIJSTEN: dict[str, list[str]] = {
    'Relatietype': ['Primair', 'Ondersteunend'],
    'Projectstatus': ['Backlog', 'Actief', 'On-hold', 'Gereed', 'Vervallen'],
    'RAG-status': ['Rood', 'Oranje', 'Groen'],
    'Org-relatie status': ['Concept', 'Gevalideerd', 'Vervallen'],
    'Actief': ['Ja', 'Nee'],
    'Product-type': ['deliverable', 'mijlpaal'],
}


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'


def _new_workbook(format_: str) -> tuple[Workbook, dict[str, Worksheet]]:
    wb = Workbook()
    wb.remove(wb.active)
    headers_map = OUD_SHEET_HEADERS if format_ == 'oud' else NIEUW_SHEET_HEADERS
    order = OUD_SHEET_ORDER if format_ == 'oud' else NIEUW_SHEET_ORDER
    sheets: dict[str, Worksheet] = {}
    for name in order:
        ws = create_safe_sheet(wb, name)
        _write_header(ws, headers_map[name])
        sheets[name] = ws
    return wb, sheets


def _write_configuratie(wb: Workbook, meta: dict[str, Any], format_label: str, mode_label: str) -> None:
    ws = create_safe_sheet(wb, CONFIG_SHEET_NAME)
    ws.append(['Sleutel', 'Waarde'])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    rows = [
        ('Doelenboom', meta.get('doelenboom', '')),
        ('Tenant', meta.get('tenant', '')),
        ('Formaat', format_label),
        ('Modus', mode_label),
        ('Export-formaatversie', TREE_EXPORT_FORMAT_VERSION),
        ('Geëxporteerd op', meta.get('exportedAt', '')),
        ('Geëxporteerd door', meta.get('exportedBy', '')),
        ('Bron', 'Doelenboom platform'),
    ]
    for r in rows:
        ws.append(list(r))
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 44


def _write_kolommen(wb: Workbook, columns: list[dict[str, Any]]) -> None:
    """Schrijft een leesbare 'Kolommen'-tab met de volledige kolomconfiguratie
    van deze doelenboom (zie docs/kolommen-configuratie-ontwerp.md) — puur
    documentair, net als de Configuratie-tab: hiermee is uit het losse
    Excel-bestand zelf af te lezen welke typen/kolommen deze doelenboom op het
    moment van exporteren had, zonder dat je in de app hoeft te kijken. Wordt
    voor beide formaten (oud/nieuw) en beide modi (template/data) geschreven,
    ook al is de kolomconfiguratie bij het "oud" formaat altijd de 8
    standaardkolommen (zie is_standard_columns) — ook dan is het nuttig om
    bevestigd te zien."""
    ws = create_safe_sheet(wb, KOLOMMEN_SHEET_NAME)
    headers = [
        'Volgorde', 'Type', 'Titel', 'Ondertitel', 'Kleur', 'Smal',
        'Projectrol', 'Label naar volgende kolom', 'Lettergrootte knoop',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'
    ordered = sorted(columns, key=lambda c: c.get('position', 0))
    for c in ordered:
        node_font_size = c.get('nodeFontSize')
        ws.append([
            c.get('position', ''),
            c.get('typeName', ''),
            c.get('title', ''),
            c.get('subtitle', ''),
            c.get('color', ''),
            'Ja' if c.get('isNarrow') else 'Nee',
            'Ja' if c.get('isProjectRole') else 'Nee',
            c.get('relationLabelToNext') or '',
            node_font_size if node_font_size is not None else '',
        ])
    widths = [10, 22, 22, 22, 10, 8, 11, 26, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_validatielijsten(wb: Workbook, columns: list[dict[str, Any]]) -> dict[str, str]:
    """Schrijft de _Validatielijsten-tab en geeft per lijst de celrange terug
    (bv. 'Type' -> '_Validatielijsten!$A$2:$A$9') voor gebruik in Data Validation.
    'Type' komt (i.t.t. de andere lijsten) niet uit de vaste VALIDATIELIJSTEN-
    constante maar dynamisch uit de kolomconfiguratie van deze doelenboom."""
    lijsten: dict[str, list[str]] = {'Type': _type_names_from_columns(columns), **VALIDATIELIJSTEN}
    ws = create_safe_sheet(wb, VALIDATIELIJSTEN_SHEET_NAME)
    headers = list(lijsten.keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    max_len = max(len(v) for v in lijsten.values())
    for i in range(max_len):
        row = [lijsten[key][i] if i < len(lijsten[key]) else None for key in headers]
        ws.append(row)

    ranges: dict[str, str] = {}
    for idx, key in enumerate(headers, start=1):
        col = get_column_letter(idx)
        n = len(lijsten[key])
        ranges[key] = f'{VALIDATIELIJSTEN_SHEET_NAME}!${col}$2:${col}${n + 1}'
    return ranges


def _add_dropdown(ws: Worksheet, formula1: str, cell_range: str, allow_blank: bool = True) -> None:
    dv = DataValidation(type='list', formula1=formula1, allow_blank=allow_blank, showErrorMessage=True)
    dv.error = 'Kies een waarde uit de lijst.'
    dv.errorTitle = 'Ongeldige waarde'
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _apply_data_validation(sheets: dict[str, Worksheet], ranges: dict[str, str]) -> None:
    ref_ws = sheets['Referentietabel']
    _add_dropdown(ref_ws, ranges['Type'], 'B2:B10000', allow_blank=False)
    _add_dropdown(ref_ws, ranges['Actief'], 'I2:I10000', allow_blank=False)

    rel_ws = sheets['Relaties']
    _add_dropdown(rel_ws, ranges['Relatietype'], 'C2:C10000')

    proj_ws = sheets['Projecten']
    _add_dropdown(proj_ws, ranges['Projectstatus'], 'D2:D10000')
    _add_dropdown(proj_ws, ranges['RAG-status'], 'E2:E10000')

    prod_ws = sheets['Producten']
    _add_dropdown(prod_ws, ranges['Product-type'], 'E2:E10000')

    act_ws = sheets['Activiteiten']
    _add_dropdown(act_ws, ranges['Actief'], 'H2:H10000')
    _add_dropdown(act_ws, ranges['Actief'], 'I2:I10000')

    obo_ws = sheets['OB-Organisatie relaties']
    _add_dropdown(obo_ws, ranges['Relatietype'], 'E2:E10000')
    _add_dropdown(obo_ws, ranges['Org-relatie status'], 'G2:G10000')


def _relatietype_label(weight: str | None) -> str:
    if weight == 'primair':
        return 'Primair'
    if weight == 'ondersteunend':
        return 'Ondersteunend'
    return ''


def build_template_workbook(
    format_: str, meta: dict[str, Any] | None = None, columns: list[dict[str, Any]] | None = None
) -> bytes:
    meta = meta or {}
    columns = columns or []
    wb, sheets = _new_workbook(format_)
    if format_ == 'nieuw':
        ranges = _write_validatielijsten(wb, columns)
        _apply_data_validation(sheets, ranges)
    _write_kolommen(wb, columns)
    _write_configuratie(wb, meta, 'Nieuw' if format_ == 'nieuw' else 'Oud', 'Lege template')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_products(sheets: dict[str, Worksheet], tree: dict[str, Any], by_code: dict[str, Any]) -> None:
    """Vult de Producten-tab — identiek voor 'oud' en 'nieuw', daarom hier één
    keer, aangeroepen door zowel _fill_oud als _fill_nieuw (mirroring
    _fill_activities hieronder). tree['products']/tree['productDependencies']
    komen rechtstreeks uit fetchTree() (api/src/routes/tree.ts), al per
    projectcode gegroepeerd. Vóór deze functie had de Producten-tab geen
    'Hangt af van'-kolom: productafhankelijkheden gingen dus verloren bij een
    "boom leegmaken + uit de volledige-boom-Excel reconstrueren"-scenario,
    ook al exporteerde/importeerde de aparte één-project-export
    (project_workbook.py) ze wel correct."""
    prod_ws = sheets['Producten']
    product_name_by_id: dict[Any, str] = {}
    for products in (tree.get('products', {}) or {}).values():
        for p in products:
            product_name_by_id[p.get('id')] = p.get('name', '')

    product_preds: dict[Any, list[tuple[str, int | float, str]]] = {}
    for deps in (tree.get('productDependencies', {}) or {}).values():
        for d in deps:
            product_preds.setdefault(d.get('successorId'), []).append((
                product_name_by_id.get(d.get('predecessorId'), '?'),
                d.get('lagAmount') or 0,
                d.get('lagEenheid') or 'd',
            ))

    for project_code, products in (tree.get('products', {}) or {}).items():
        project_el = by_code.get(project_code)
        project_name = project_el['name'] if project_el else ''
        for p in products:
            preds = product_preds.get(p.get('id'), [])
            prod_ws.append([
                p.get('code', ''), project_code, project_name, p.get('name', ''),
                p.get('type', 'deliverable'), p.get('omschrijving', ''), p.get('pctGereed', 0),
                p.get('verwachteDatum') or '', p.get('werkelijkeDatum') or '', p.get('opmerking', ''),
                '; '.join(format_product_dependency(n, la, le) for (n, la, le) in preds),
            ])


def _fill_activities(sheets: dict[str, Worksheet], tree: dict[str, Any], by_code: dict[str, Any]) -> None:
    """Vult de Activiteiten-tab — identiek voor 'oud' en 'nieuw' (net als
    Producten hierboven), daarom hier één keer, aangeroepen door zowel
    _fill_oud als _fill_nieuw. tree['activities']/tree['dependencies'] komen
    rechtstreeks uit fetchTree() (api/src/routes/tree.ts), al per projectcode
    gegroepeerd — zelfde vorm als tree['products']/tree['productDependencies']
    hierboven. Vóór deze functie ontbrak een Activiteiten-tab in de volledige-
    boom-export volledig: activiteiten waren dan alleen via de aparte
    één-project-export (project_workbook.py) terug te halen, dus een
    "boom leegmaken + uit Excel reconstrueren"-scenario zou ze kwijtraken als
    niet ook voor elk project apart geëxporteerd was."""
    act_ws = sheets['Activiteiten']
    activity_name_by_id: dict[Any, str] = {}
    for acts in (tree.get('activities', {}) or {}).values():
        for a in acts:
            activity_name_by_id[a.get('id')] = a.get('name', '')

    activity_preds: dict[Any, list[tuple[str, str, int]]] = {}
    for deps in (tree.get('dependencies', {}) or {}).values():
        for d in deps:
            activity_preds.setdefault(d.get('successorId'), []).append((
                activity_name_by_id.get(d.get('predecessorId'), '?'),
                d.get('type') or 'FS',
                d.get('lagDays') or 0,
            ))

    for project_code, acts in (tree.get('activities', {}) or {}).items():
        project_el = by_code.get(project_code)
        project_name = project_el['name'] if project_el else ''
        for a in acts:
            preds = activity_preds.get(a.get('id'), [])
            act_ws.append([
                a.get('id'), project_code, project_name, a.get('name', ''),
                a.get('startDate') or '', a.get('endDate') or '', a.get('omschrijving', ''),
                'Ja' if a.get('isMilestone') else 'Nee',
                'Ja' if a.get('isSummary') else 'Nee',
                '; '.join(format_dependency(n, t, l) for (n, t, l) in preds),
            ])


def _fill_oud(sheets: dict[str, Worksheet], tree: dict[str, Any]) -> None:
    elements = tree.get('elements', [])
    by_code = {el['code']: el for el in elements}
    project_status = tree.get('projectStatus', {}) or {}

    ws = sheets['Referentietabel']
    for el in elements:
        ps = project_status.get(el['code'], {})
        ws.append([
            el.get('code', ''), el.get('type', ''), el.get('name', ''),
            el.get('description', ''), el.get('parent_text', ''),
            el.get('kpi', ''), el.get('taakveld', ''), el.get('subtaakveld', ''),
            ps.get('projectstatus', ''), ps.get('rag', ''), ps.get('toelichting', ''),
            ps.get('gerapporteerdOp') or '',
        ])

    cap_ob_ws = sheets['Capability-OB relaties']
    proj_cap_ws = sheets['Project-Capability relaties']
    for e in tree.get('edges', []):
        src = by_code.get(e.get('source'))
        tgt = by_code.get(e.get('target'))
        if not src or not tgt:
            continue
        rel = _relatietype_label(e.get('weight'))
        if src.get('type') == 'Capability' and tgt.get('type') == 'Operationele benefit':
            cap_ob_ws.append([src['code'], src['name'], tgt['code'], tgt['name'], rel, e.get('toelichting', '')])
        elif src.get('type') == 'Project' and tgt.get('type') == 'Capability':
            proj_cap_ws.append([src['code'], src['name'], tgt['code'], tgt['name'], rel, e.get('toelichting', '')])
        # Overige (verticale) edges komen niet uit een apart relatietabblad — die zitten
        # al in "Bovenliggend element" op de Referentietabel-rij van het onderliggende element.

    proj_ws = sheets['Projecten']
    for el in elements:
        if el.get('type') != 'Project':
            continue
        ps = project_status.get(el['code'], {})
        proj_ws.append([
            el['code'], el['name'], ps.get('clusterPpt', ''), ps.get('projectstatus', ''),
            ps.get('rag', ''), ps.get('toelichting', ''), ps.get('gerapporteerdOp') or '', el.get('description', ''),
        ])

    _fill_products(sheets, tree, by_code)
    _fill_activities(sheets, tree, by_code)

    tags_ws = sheets['Tags']
    for t in tree.get('tags', []):
        tags_ws.append([t.get('code', ''), t.get('name', ''), t.get('categorie', ''), t.get('omschrijving', '')])

    et_ws = sheets['Element-Tag relaties']
    tag_by_code = {t['code']: t for t in tree.get('tags', [])}
    for element_code, tag_codes in (tree.get('elementTags', {}) or {}).items():
        el = by_code.get(element_code)
        if not el:
            continue
        for tag_code in tag_codes:
            tag = tag_by_code.get(tag_code)
            et_ws.append([el['code'], el['type'], el['name'], tag_code, tag['name'] if tag else '', ''])

    org_ws = sheets['Organisatieonderdelen']
    for o in tree.get('orgUnits', []):
        org_ws.append([o.get('code', ''), o.get('name', ''), o.get('omschrijving', '')])

    obo_ws = sheets['OB-Organisatie relaties']
    org_by_code = {o['code']: o for o in tree.get('orgUnits', [])}
    for ob_code, rels in (tree.get('obOrg', {}) or {}).items():
        ob_el = by_code.get(ob_code)
        if not ob_el:
            continue
        for r in rels:
            org = org_by_code.get(r.get('org'))
            obo_ws.append([
                ob_code, ob_el['name'], r.get('org', ''), org['name'] if org else '',
                r.get('relatietype', ''), r.get('toelichting', ''), r.get('status', ''),
            ])


def _fill_nieuw(sheets: dict[str, Worksheet], tree: dict[str, Any]) -> None:
    elements = tree.get('elements', [])
    by_code = {el['code']: el for el in elements}
    project_status = tree.get('projectStatus', {}) or {}

    # Referentietabel: pure stamdata + Volgorde + Actief (§3/§5/§6 van het voorstel).
    # Alles wat momenteel in de database staat is per definitie actief — "historisch"-
    # rijen worden al bij import weggefilterd en nooit opgeslagen (zie parser.py).
    ref_ws = sheets['Referentietabel']
    for el in elements:
        ref_ws.append([
            el.get('code', ''), el.get('type', ''), el.get('name', ''), el.get('description', ''),
            el.get('kpi', ''), el.get('taakveld', ''), el.get('subtaakveld', ''),
            el.get('sort_order', 0), 'Ja',
        ])

    # Relaties: één generieke tab i.p.v. Capability-OB/Project-Capability/vrije tekst (§2).
    rel_ws = sheets['Relaties']
    for e in tree.get('edges', []):
        rel_ws.append([e.get('source', ''), e.get('target', ''), _relatietype_label(e.get('weight')), e.get('toelichting', '')])

    proj_ws = sheets['Projecten']
    for el in elements:
        if el.get('type') != 'Project':
            continue
        ps = project_status.get(el['code'], {})
        proj_ws.append([
            el['code'], el['name'], ps.get('clusterPpt', ''), ps.get('projectstatus', ''),
            ps.get('rag', ''), ps.get('toelichting', ''), ps.get('gerapporteerdOp') or '', el.get('description', ''),
        ])

    _fill_products(sheets, tree, by_code)
    _fill_activities(sheets, tree, by_code)

    tags_ws = sheets['Tags']
    for t in tree.get('tags', []):
        tags_ws.append([t.get('code', ''), t.get('name', ''), t.get('categorie', ''), t.get('omschrijving', '')])

    et_ws = sheets['Element-Tag relaties']
    tag_by_code = {t['code']: t for t in tree.get('tags', [])}
    for element_code, tag_codes in (tree.get('elementTags', {}) or {}).items():
        el = by_code.get(element_code)
        if not el:
            continue
        for tag_code in tag_codes:
            tag = tag_by_code.get(tag_code)
            et_ws.append([el['code'], el['type'], el['name'], tag_code, tag['name'] if tag else '', ''])

    org_ws = sheets['Organisatieonderdelen']
    for o in tree.get('orgUnits', []):
        org_ws.append([o.get('code', ''), o.get('name', ''), o.get('omschrijving', '')])

    obo_ws = sheets['OB-Organisatie relaties']
    org_by_code = {o['code']: o for o in tree.get('orgUnits', [])}
    for ob_code, rels in (tree.get('obOrg', {}) or {}).items():
        ob_el = by_code.get(ob_code)
        if not ob_el:
            continue
        for r in rels:
            org = org_by_code.get(r.get('org'))
            obo_ws.append([
                ob_code, ob_el['name'], r.get('org', ''), org['name'] if org else '',
                r.get('relatietype', ''), r.get('toelichting', ''), r.get('status', ''),
            ])


def build_data_workbook(
    format_: str, tree: dict[str, Any], meta: dict[str, Any] | None = None, columns: list[dict[str, Any]] | None = None
) -> bytes:
    meta = meta or {}
    columns = columns if columns is not None else (tree.get('columns') or [])
    wb, sheets = _new_workbook(format_)
    if format_ == 'oud':
        _fill_oud(sheets, tree)
    else:
        _fill_nieuw(sheets, tree)
        ranges = _write_validatielijsten(wb, columns)
        _apply_data_validation(sheets, ranges)
    _write_kolommen(wb, columns)
    _write_configuratie(wb, meta, 'Nieuw' if format_ == 'nieuw' else 'Oud', 'Met huidige data')
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
