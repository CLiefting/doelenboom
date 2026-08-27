"""Tests voor app/project_workbook.py — bouwt een Project-Excel-bestand voor
één project en laadt het weer in via parse_project_workbook, om te
controleren dat alle velden (incl. afhankelijkheden, tags/organisatie-
onderdelen) round-trip-veilig zijn."""
from __future__ import annotations

from app.project_workbook import (
    ACTIVITIES_SHEET, PRODUCTS_SHEET, PROJECT_SHEET,
    _format_dependency, _parse_dependency_entry,
    build_project_workbook, parse_project_workbook,
)


def make_data():
    return {
        'project': {
            'code': 'NP37', 'name': 'Sweepen', 'description': 'Test-omschrijving',
            'status': {
                'projectstatus': 'actief', 'rag': 'groen', 'toelichting': 'Op schema',
                'gerapporteerdOp': '2026-08-01', 'clusterPpt': 'Cluster A',
            },
            'tags': ['IGO', 'Innovatie'],
            'orgs': [{'name': 'HRB-S', 'relatietype': 'Primair'}, {'name': 'BSB', 'relatietype': 'Ondersteunend'}],
        },
        'products': [
            {
                'id': 1, 'name': 'PID', 'type': 'deliverable', 'omschrijving': 'Projectinitiatiedocument',
                'pctGereed': 100, 'verwachteDatum': '2026-09-15', 'werkelijkeDatum': '2026-09-10',
                'deadline': None, 'duur': None, 'duurEenheid': 'd', 'businessValue': None, 'opmerking': '',
            },
            {
                'id': 2, 'name': 'Adviesrapport', 'type': 'deliverable', 'omschrijving': '',
                'pctGereed': 30, 'verwachteDatum': '2026-08-29', 'werkelijkeDatum': None,
                'deadline': '2026-10-01', 'duur': 10, 'duurEenheid': 'm', 'businessValue': 100, 'opmerking': 'let op',
            },
            {
                'id': 3, 'name': 'GO/NO-GO', 'type': 'mijlpaal', 'omschrijving': '',
                'pctGereed': 0, 'verwachteDatum': '2027-04-01', 'werkelijkeDatum': None,
                'deadline': None, 'duur': None, 'duurEenheid': 'd', 'businessValue': None, 'opmerking': '',
            },
        ],
        'productDependencies': [
            {'id': 11, 'predecessorId': 1, 'successorId': 2},
            {'id': 12, 'predecessorId': 2, 'successorId': 3},
        ],
        'activities': [
            {'id': 101, 'name': 'Taak A', 'startDate': '2026-08-01', 'endDate': '2026-08-10',
             'omschrijving': 'Eerste taak', 'isMilestone': False, 'isSummary': False},
            {'id': 102, 'name': 'Taak B', 'startDate': '2026-08-11', 'endDate': '2026-08-11',
             'omschrijving': '', 'isMilestone': True, 'isSummary': False},
        ],
        'activityDependencies': [
            {'id': 21, 'predecessorId': 101, 'successorId': 102, 'type': 'FS', 'lagDays': 0},
        ],
    }


def make_meta():
    return {'doelenboom': 'FPBB', 'tenant': 'Voorbeeld', 'exportedAt': '2026-08-27T19:00:00Z', 'exportedBy': 'test@example.com'}


class TestBuildProjectWorkbook:
    def test_bevat_alle_tabbladen(self):
        content = build_project_workbook(make_data(), make_meta())
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(content))
        assert set(wb.sheetnames) >= {'Info', PROJECT_SHEET, PRODUCTS_SHEET, ACTIVITIES_SHEET}

    def test_producten_tab_heeft_hangt_af_van_kolom_gevuld(self):
        content = build_project_workbook(make_data(), make_meta())
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(content))
        ws = wb[PRODUCTS_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        idx = header.index('Hangt af van')
        by_name = {r[1]: r[idx] for r in rows[1:]}
        assert by_name['Adviesrapport'] == 'PID'
        assert by_name['GO/NO-GO'] == 'Adviesrapport'
        # openpyxl geeft een lege cel terug als None, niet als '' (ook al is
        # '' geschreven) — build_project_workbook schrijft hier terecht een
        # lege string (zie _split_entries/clean_text die dat ook zo verwachten).
        assert by_name['PID'] is None

    def test_activiteiten_tab_heeft_voorgangers_kolom(self):
        content = build_project_workbook(make_data(), make_meta())
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(content))
        ws = wb[ACTIVITIES_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        idx = header.index('Voorgangers')
        by_name = {r[1]: r[idx] for r in rows[1:]}
        assert by_name['Taak B'] == 'Taak A'  # FS + lag 0 -> geen suffix
        assert by_name['Taak A'] is None  # zie toelichting bij PID hierboven


class TestDependencyFormatting:
    def test_fs_zonder_vertraging_blijft_kaal(self):
        assert _format_dependency('Taak A', 'FS', 0) == 'Taak A'

    def test_ander_type_of_vertraging_krijgt_suffix(self):
        assert _format_dependency('Taak B', 'SS', 0) == 'Taak B (SS)'
        assert _format_dependency('Taak C', 'FS', 2) == 'Taak C (FS+2)'
        assert _format_dependency('Taak D', 'FF', -1) == 'Taak D (FF-1)'

    def test_parse_is_inverse_van_format(self):
        for name, dep_type, lag in [('Taak A', 'FS', 0), ('Taak B', 'SS', 0), ('Taak C', 'FS', 2), ('Taak D', 'FF', -1)]:
            formatted = _format_dependency(name, dep_type, lag)
            parsed_name, parsed_type, parsed_lag = _parse_dependency_entry(formatted)
            assert (parsed_name, parsed_type, parsed_lag) == (name, dep_type, lag)


class TestRoundTrip:
    def test_volledige_roundtrip(self):
        content = build_project_workbook(make_data(), make_meta())
        status, report, parsed = parse_project_workbook(content)
        assert status == 'ok', report
        assert not report['errors']
        assert not report['warnings']

        assert parsed['project']['code'] == 'NP37'
        assert parsed['project']['name'] == 'Sweepen'
        assert parsed['project']['projectstatus'] == 'actief'
        assert parsed['project']['rag'] == 'groen'
        assert parsed['project']['toelichting'] == 'Op schema'
        assert parsed['project']['gerapporteerdOp'] == '2026-08-01'
        assert parsed['project']['clusterPpt'] == 'Cluster A'
        assert parsed['project']['tags'] == ['IGO', 'Innovatie']
        assert parsed['project']['orgs'] == [
            {'name': 'HRB-S', 'relatietype': 'Primair'},
            {'name': 'BSB', 'relatietype': 'Ondersteunend'},
        ]

        assert len(parsed['products']) == 3
        by_name = {p['name']: p for p in parsed['products']}
        assert by_name['PID']['id'] == 1
        assert by_name['PID']['pctGereed'] == 100
        assert by_name['PID']['werkelijkeDatum'] == '2026-09-10'
        assert by_name['Adviesrapport']['duur'] == 10
        assert by_name['Adviesrapport']['duurEenheid'] == 'm'
        assert by_name['Adviesrapport']['businessValue'] == 100
        assert by_name['Adviesrapport']['deadline'] == '2026-10-01'
        assert by_name['Adviesrapport']['dependsOnNames'] == ['PID']
        assert by_name['GO/NO-GO']['type'] == 'mijlpaal'
        assert by_name['GO/NO-GO']['dependsOnNames'] == ['Adviesrapport']
        assert by_name['PID']['dependsOnNames'] == []

        assert len(parsed['activities']) == 2
        act_by_name = {a['name']: a for a in parsed['activities']}
        assert act_by_name['Taak A']['id'] == 101
        assert act_by_name['Taak A']['isMilestone'] is False
        assert act_by_name['Taak B']['isMilestone'] is True
        assert act_by_name['Taak B']['predecessors'] == [{'name': 'Taak A', 'type': 'FS', 'lagDays': 0}]
        assert act_by_name['Taak A']['predecessors'] == []

    def test_nieuwe_rij_zonder_id(self):
        data = make_data()
        data['products'].append({
            'id': 4, 'name': 'Extra deliverable', 'type': 'deliverable', 'omschrijving': '',
            'pctGereed': 0, 'verwachteDatum': None, 'werkelijkeDatum': None,
            'deadline': None, 'duur': None, 'duurEenheid': 'd', 'businessValue': None, 'opmerking': '',
        })
        content = build_project_workbook(data, make_meta())
        # Simuleer een gebruiker die de ID-cel van een nieuwe rij leeghaalt
        # door 'm handmatig als nieuwe rij toe te voegen i.p.v. de geëxporteerde
        # rij te bewerken (hier: gewoon een cel legen na het opnieuw inlezen).
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(content))
        ws = wb[PRODUCTS_SHEET]
        for row in ws.iter_rows(min_row=2):
            if row[1].value == 'Extra deliverable':
                row[0].value = None
        bio = io.BytesIO()
        wb.save(bio)
        status, report, parsed = parse_project_workbook(bio.getvalue())
        assert status == 'ok', report
        extra = next(p for p in parsed['products'] if p['name'] == 'Extra deliverable')
        assert extra['id'] is None

    def test_missing_sheets_faalt_netjes(self):
        from openpyxl import Workbook
        import io
        wb = Workbook()
        bio = io.BytesIO()
        wb.save(bio)
        status, report, parsed = parse_project_workbook(bio.getvalue())
        assert status == 'failed'
        assert parsed is None
        assert report['sheetsMissing'] == [PROJECT_SHEET, PRODUCTS_SHEET, ACTIVITIES_SHEET]

    def test_onbekende_projectstatus_geeft_warning(self):
        data = make_data()
        data['project']['status']['projectstatus'] = 'onzin-status'
        content = build_project_workbook(data, make_meta())
        status, report, parsed = parse_project_workbook(content)
        assert status == 'warning'
        assert any('Projectstatus' in w for w in report['warnings'])
        assert parsed['project']['projectstatus'] == ''
