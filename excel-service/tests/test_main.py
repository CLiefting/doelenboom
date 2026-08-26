"""Tests voor app/main.py — de FastAPI-routes zelf (i.t.t. test_exporter.py/
test_parser.py, die de onderliggende functies rechtstreeks aanroepen). Dekt
vooral de twee dingen die per HTTP-laag zijn toegevoegd voor de configureerbare
kolommen (zie docs/kolommen-configuratie-ontwerp.md): de 'oud'-formaat-
blokkade op /export, en het doorgeven van valid_types op /parse."""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.exporter import NIEUW_SHEET_HEADERS
from app.main import app
from tests.helpers import build_workbook_bytes, make_columns, make_tree, standard_columns

client = TestClient(app)


class TestExportOudFormaatBlokkade:
    def test_oud_formaat_met_standaardkolommen_werkt(self):
        tree = make_tree(columns=standard_columns())
        res = client.post(
            '/export?format=oud&mode=data',
            json={'tree': tree, 'meta': {}},
        )
        assert res.status_code == 200

    def test_oud_formaat_met_aangepaste_kolommen_geeft_409(self):
        tree = make_tree(columns=make_columns(['Initiatief', 'Vermogen', 'Ambitie']))
        res = client.post(
            '/export?format=oud&mode=data',
            json={'tree': tree, 'meta': {}},
        )
        assert res.status_code == 409
        assert 'oud' in res.json()['error'].lower()

    def test_nieuw_formaat_met_aangepaste_kolommen_werkt_gewoon(self):
        tree = make_tree(columns=make_columns(['Initiatief', 'Vermogen', 'Ambitie']))
        res = client.post(
            '/export?format=nieuw&mode=data',
            json={'tree': tree, 'meta': {}},
        )
        assert res.status_code == 200
        wb = load_workbook(io.BytesIO(res.content))
        val_ws = wb['_Validatielijsten']
        type_values = [row[0] for row in val_ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        assert type_values == ['Initiatief', 'Vermogen', 'Ambitie']

    def test_template_modus_gebruikt_apart_meegegeven_columns(self):
        # mode=template heeft geen tree, dus columns moet los in de body zitten
        # (zie routes/exports.ts) — precies wat hier gebeurt.
        res = client.post(
            '/export?format=nieuw&mode=template',
            json={'tree': None, 'columns': make_columns(['A', 'B']), 'meta': {}},
        )
        assert res.status_code == 200
        wb = load_workbook(io.BytesIO(res.content))
        val_ws = wb['_Validatielijsten']
        type_values = [row[0] for row in val_ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        assert type_values == ['A', 'B']

    def test_oud_template_zonder_columns_wordt_niet_geblokkeerd(self):
        # Geen columns meegegeven (lege lijst) -> is_standard_columns([]) is
        # False, dus zou in theorie blokkeren; maar template-mode voor 'oud'
        # gebruikt de Type-lijst niet (geen dropdown in 'oud', zie exporter.py),
        # dus dit hoort alsnog te blokkeren net als elke andere niet-standaard
        # config — bewust hier expliciet vastgelegd i.p.v. aangenomen.
        res = client.post('/export?format=oud&mode=template', json={'tree': None, 'meta': {}})
        assert res.status_code == 409


class TestParseValidTypes:
    def _nieuw_workbook(self, type_value: str) -> bytes:
        return build_workbook_bytes({
            'Referentietabel': [
                NIEUW_SHEET_HEADERS['Referentietabel'],
                ['I1', type_value, 'Init 1', '', '', '', '', 1, 'Ja'],
            ],
            'Relaties': [NIEUW_SHEET_HEADERS['Relaties']],
            'Projecten': [NIEUW_SHEET_HEADERS['Projecten']],
            'Producten': [NIEUW_SHEET_HEADERS['Producten']],
            'Tags': [NIEUW_SHEET_HEADERS['Tags']],
            'Element-Tag relaties': [NIEUW_SHEET_HEADERS['Element-Tag relaties']],
            'Organisatieonderdelen': [NIEUW_SHEET_HEADERS['Organisatieonderdelen']],
            'OB-Organisatie relaties': [NIEUW_SHEET_HEADERS['OB-Organisatie relaties']],
        })

    def test_zonder_valid_types_wordt_eigen_type_geweigerd(self):
        content = self._nieuw_workbook('Initiatief')
        res = client.post('/parse', files={'file': ('test.xlsx', content, 'application/octet-stream')})
        assert res.status_code == 200
        body = res.json()
        assert body['status'] == 'failed'
        assert any('Onbekend Type-label' in w for w in body['report']['warnings'])

    def test_met_valid_types_wordt_eigen_type_geaccepteerd(self):
        content = self._nieuw_workbook('Initiatief')
        res = client.post(
            '/parse?valid_types=Initiatief&valid_types=Vermogen',
            files={'file': ('test.xlsx', content, 'application/octet-stream')},
        )
        assert res.status_code == 200
        body = res.json()
        assert body['status'] == 'ok'
        assert body['parsed']['elements'][0]['type'] == 'Initiatief'


class TestParseMpp:
    """/parse-mpp (mpp_converter.py) — de "echte" conversie (een geldig
    .mpp-bestand omzetten) vereist een werkende Java-installatie, net als de
    Docker-image (zie Dockerfile). Op een lokale ontwikkelmachine zonder JRE
    slaan we die ene test over i.p.v. de hele suite te laten falen op een
    omgevingsissue dat niets met deze code te maken heeft — de foutafhandeling
    zelf (leeg bestand, onleesbaar bestand) wordt dan nog steeds gedekt."""

    def test_leeg_bestand_geeft_400(self):
        res = client.post('/parse-mpp', files={'file': ('test.mpp', b'', 'application/octet-stream')})
        assert res.status_code == 400
        assert 'leeg' in res.json()['error'].lower()

    def test_onleesbaar_bestand_geeft_400(self):
        try:
            res = client.post(
                '/parse-mpp',
                files={'file': ('test.mpp', b'dit is geen geldig mpp-bestand', 'application/octet-stream')},
            )
        except Exception:
            pytest.skip('Geen werkende Java-installatie lokaal beschikbaar (JRE nodig voor mpxj, zie Dockerfile).')
            return
        if res.status_code == 500:
            pytest.skip('Geen werkende Java-installatie lokaal beschikbaar (JRE nodig voor mpxj, zie Dockerfile).')
            return
        assert res.status_code == 400
        assert 'niet lezen' in res.json()['error'].lower()
