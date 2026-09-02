"""Tests voor app/exporter.py — bouwt template/data-workbooks en laadt ze terug
in openpyxl om headers, datarijen en (voor het nieuwe formaat) de dropdown-
validatie te controleren."""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from app.exporter import (
    NIEUW_SHEET_HEADERS, OUD_SHEET_HEADERS, STANDARD_TYPE_NAMES,
    build_data_workbook, build_template_workbook, is_standard_columns,
)
from tests.helpers import make_columns, make_tree, standard_columns


def load(content: bytes):
    return load_workbook(io.BytesIO(content))


class TestTemplateWorkbook:
    def test_oud_template_heeft_alle_tabbladen_met_alleen_headers(self):
        wb = load(build_template_workbook('oud'))
        for name, headers in OUD_SHEET_HEADERS.items():
            ws = wb[name]
            assert [c.value for c in ws[1]] == headers
            assert ws.max_row == 1  # geen datarijen
        assert 'Configuratie' in wb.sheetnames

    def test_nieuw_template_heeft_validatielijsten_en_dropdowns(self):
        wb = load(build_template_workbook('nieuw'))
        for name, headers in NIEUW_SHEET_HEADERS.items():
            assert [c.value for c in wb[name][1]] == headers
        assert '_Validatielijsten' in wb.sheetnames

        ref_ws = wb['Referentietabel']
        # De Type-dropdown staat op kolom B (B2:B10000, zie exporter.py
        # _apply_data_validation) en verwijst naar de _Validatielijsten-tab —
        # de dropdown zelf draagt de tekst "Type" niet, dus zoeken op de
        # celrange i.p.v. op de formuletekst.
        type_dvs = [dv for dv in ref_ws.data_validations.dataValidation if 'B2:B10000' in str(dv.sqref)]
        assert type_dvs, 'verwacht een Type-dropdown op de Referentietabel'
        assert '_Validatielijsten' in type_dvs[0].formula1

    def test_configuratie_tab_bevat_meta(self):
        wb = load(build_template_workbook('oud', {'doelenboom': 'Mijn boom', 'tenant': 'Mijn tenant'}))
        rows = {row[0].value: row[1].value for row in wb['Configuratie'].iter_rows(min_row=2)}
        assert rows['Doelenboom'] == 'Mijn boom'
        assert rows['Tenant'] == 'Mijn tenant'
        assert rows['Formaat'] == 'Oud'
        assert rows['Modus'] == 'Lege template'
        assert rows['Export-formaatversie']


class TestDataWorkbookOud:
    def test_referentietabel_bevat_elementrijen(self):
        tree = make_tree()
        wb = load(build_data_workbook('oud', tree))
        ws = wb['Referentietabel']
        codes = [row[0].value for row in ws.iter_rows(min_row=2)]
        assert codes == ['OB1', 'C1', 'P1']

    def test_edges_verdeeld_over_de_twee_relatietabbladen(self):
        tree = make_tree()
        wb = load(build_data_workbook('oud', tree))
        cap_ob = [tuple(c.value for c in row) for row in wb['Capability-OB relaties'].iter_rows(min_row=2)]
        proj_cap = [tuple(c.value for c in row) for row in wb['Project-Capability relaties'].iter_rows(min_row=2)]
        assert cap_ob == [('C1', 'Capability 1', 'OB1', 'OB 1', 'Primair', 'Waarom C1->OB1')]
        # Een leeg-tekst-cel ('') komt bij het opnieuw inladen van het .xlsx-
        # bestand terug als None (openpyxl-gedrag, geen bug) — parser.py's
        # clean_text() behandelt beide gelijk, dus dat is verderop onschadelijk.
        assert proj_cap == [('P1', 'Project 1', 'C1', 'Capability 1', 'Ondersteunend', None)]

    def test_producten_bevatten_type_kolom(self):
        tree = make_tree()
        wb = load(build_data_workbook('oud', tree))
        rows = [tuple(c.value for c in row) for row in wb['Producten'].iter_rows(min_row=2)]
        assert rows[0][4] == 'deliverable'
        assert rows[1][4] == 'mijlpaal'

    def test_projecten_tab_bevat_status_en_cluster(self):
        tree = make_tree()
        wb = load(build_data_workbook('oud', tree))
        row = next(wb['Projecten'].iter_rows(min_row=2, values_only=True))
        assert row[0] == 'P1'
        assert row[2] == 'Cluster A'
        assert row[3] == 'Actief'
        assert row[4] == 'Groen'

    def test_tags_en_org_units_worden_geschreven(self):
        tree = make_tree()
        wb = load(build_data_workbook('oud', tree))
        assert next(wb['Tags'].iter_rows(min_row=2, values_only=True))[:2] == ('T1', 'Tag 1')
        assert next(wb['Organisatieonderdelen'].iter_rows(min_row=2, values_only=True))[:2] == ('O1', 'Org-unit 1')


class TestDataWorkbookNieuw:
    def test_relaties_tab_bevat_alle_edges_in_een_tabblad(self):
        tree = make_tree()
        wb = load(build_data_workbook('nieuw', tree))
        rows = [tuple(c.value for c in row) for row in wb['Relaties'].iter_rows(min_row=2)]
        # Zie de toelichting bij test_edges_verdeeld_over_de_twee_relatietabbladen
        # hierboven — een lege toelichting-cel komt terug als None, niet ''.
        assert rows == [('C1', 'OB1', 'Primair', 'Waarom C1->OB1'), ('P1', 'C1', 'Ondersteunend', None)]

    def test_referentietabel_heeft_volgorde_en_actief_kolom(self):
        tree = make_tree()
        wb = load(build_data_workbook('nieuw', tree))
        headers = [c.value for c in wb['Referentietabel'][1]]
        assert headers == NIEUW_SHEET_HEADERS['Referentietabel']
        row = next(wb['Referentietabel'].iter_rows(min_row=2, values_only=True))
        assert row[-1] == 'Ja'  # Actief


class TestActiviteitenTab:
    """De Activiteiten-tab is identiek gevuld voor beide formaten (net als
    Producten, zie _fill_activities in exporter.py) — daarom hier één
    testklasse i.p.v. losse oud/nieuw-varianten."""

    @pytest.mark.parametrize('format_', ['oud', 'nieuw'])
    def test_activiteiten_tab_bevat_projectkoppeling_en_voorgangers(self, format_):
        tree = make_tree()
        wb = load(build_data_workbook(format_, tree))
        rows = {row[3]: row for row in wb['Activiteiten'].iter_rows(min_row=2, values_only=True)}
        assert rows['Taak A'][1:3] == ('P1', 'Project 1')  # Project-ID, Project
        assert rows['Taak A'][4] == '2026-08-01'  # Startdatum
        assert rows['Taak A'][7] == 'Nee'  # Mijlpaal
        assert rows['Taak A'][9] is None  # Voorgangers (geen)
        assert rows['Taak B'][7] == 'Ja'  # Mijlpaal
        assert rows['Taak B'][9] == 'Taak A (FS+2)'  # Voorgangers, met type+vertraging

    @pytest.mark.parametrize('format_', ['oud', 'nieuw'])
    def test_activiteiten_tab_headers(self, format_):
        tree = make_tree()
        wb = load(build_data_workbook(format_, tree))
        headers_map = OUD_SHEET_HEADERS if format_ == 'oud' else NIEUW_SHEET_HEADERS
        assert [c.value for c in wb['Activiteiten'][1]] == headers_map['Activiteiten']

    def test_lege_activiteiten_geeft_lege_tab(self):
        tree = make_tree(activities={}, dependencies={})
        wb = load(build_data_workbook('oud', tree))
        assert wb['Activiteiten'].max_row == 1  # alleen de headerrij


class TestDynamischeTypeLijst:
    """Kolomconfiguratie (zie docs/kolommen-configuratie-ontwerp.md) i.p.v. de
    voorheen hardgecodeerde VALIDATIELIJSTEN['Type'] — de Type-kolom in
    _Validatielijsten (en dus de dropdown op de Referentietabel) moet exact de
    geconfigureerde kolommen van déze doelenboom weerspiegelen, in hun eigen
    volgorde, niet meer de 8 standaardtypes."""

    def test_custom_kolommen_in_validatielijsten_type_kolom(self):
        columns = make_columns(['Initiatief', 'Vermogen', 'Ambitie'])
        wb = load(build_template_workbook('nieuw', columns=columns))
        val_ws = wb['_Validatielijsten']
        headers = [c.value for c in val_ws[1]]
        assert headers[0] == 'Type'
        type_col_values = [row[0] for row in val_ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        assert type_col_values == ['Initiatief', 'Vermogen', 'Ambitie']

    def test_type_dropdown_range_past_zich_aan_aantal_kolommen_aan(self):
        columns = make_columns(['A', 'B'])
        wb = load(build_template_workbook('nieuw', columns=columns))
        ref_ws = wb['Referentietabel']
        type_dvs = [dv for dv in ref_ws.data_validations.dataValidation if 'B2:B10000' in str(dv.sqref)]
        assert type_dvs
        assert '$A$2:$A$3' in type_dvs[0].formula1  # 2 kolommen -> rijen 2 t/m 3

    def test_zonder_columns_valt_terug_op_standaardtypes(self):
        wb = load(build_template_workbook('nieuw'))
        val_ws = wb['_Validatielijsten']
        type_col_values = [row[0] for row in val_ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        assert type_col_values == STANDARD_TYPE_NAMES

    def test_build_data_workbook_haalt_columns_uit_tree_als_niet_apart_gegeven(self):
        tree = make_tree(columns=make_columns(['Initiatief', 'Vermogen']))
        wb = load(build_data_workbook('nieuw', tree))
        val_ws = wb['_Validatielijsten']
        type_col_values = [row[0] for row in val_ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]]
        assert type_col_values == ['Initiatief', 'Vermogen']


class TestIsStandardColumns:
    """Bepaalt of het 'oud' Excel-formaat nog aangeboden mag worden (zie
    routes/exports.ts en main.py::export) — alleen als de kolommen exact de 8
    standaardtypes zijn, in dezelfde volgorde."""

    def test_standaardkolommen_zijn_standaard(self):
        assert is_standard_columns(standard_columns()) is True

    def test_andere_volgorde_is_niet_standaard(self):
        reversed_cols = make_columns(list(reversed(STANDARD_TYPE_NAMES)))
        assert is_standard_columns(reversed_cols) is False

    def test_extra_kolom_is_niet_standaard(self):
        cols = standard_columns() + make_columns(['Extra'])
        assert is_standard_columns(cols) is False

    def test_hernoemde_kolom_is_niet_standaard(self):
        cols = make_columns(STANDARD_TYPE_NAMES[:-1] + ['Anders'])
        assert is_standard_columns(cols) is False

    def test_titel_kleur_wijzigingen_maken_niet_uit_alleen_typename_telt(self):
        cols = standard_columns()
        cols[0]['title'] = 'Een compleet andere titel'
        cols[0]['color'] = '#FF00FF'
        assert is_standard_columns(cols) is True


class TestKolommenTab:
    """Nieuwe 'Kolommen'-tab (zie exporter.py::_write_kolommen) — documenteert
    de volledige kolomconfiguratie van de doelenboom in het geëxporteerde
    bestand zelf, voor beide formaten en beide modi."""

    HEADERS = [
        'Volgorde', 'Type', 'Titel', 'Ondertitel', 'Kleur', 'Smal',
        'Projectrol', 'Label naar volgende kolom', 'Lettergrootte knoop',
    ]

    def test_template_oud_heeft_kolommen_tab_met_standaardkolommen(self):
        wb = load(build_template_workbook('oud', columns=standard_columns()))
        assert 'Kolommen' in wb.sheetnames
        ws = wb['Kolommen']
        assert [c.value for c in ws[1]] == self.HEADERS
        types = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert types == STANDARD_TYPE_NAMES

    def test_template_nieuw_heeft_kolommen_tab(self):
        columns = make_columns(['Initiatief', 'Vermogen', 'Ambitie'])
        wb = load(build_template_workbook('nieuw', columns=columns))
        ws = wb['Kolommen']
        types = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert types == ['Initiatief', 'Vermogen', 'Ambitie']

    def test_data_workbook_kolommen_tab_bevat_volledige_configuratie(self):
        columns = make_columns(['Initiatief', 'Vermogen'])
        columns[0]['title'] = 'Het initiatief'
        columns[0]['subtitle'] = 'Ondertitel 1'
        columns[0]['color'] = '#3E6FA6'
        columns[0]['isNarrow'] = True
        columns[0]['nodeFontSize'] = 12
        columns[0]['relationLabelToNext'] = 'draagt bij aan'
        tree = make_tree(columns=columns)
        wb = load(build_data_workbook('nieuw', tree))
        ws = wb['Kolommen']
        rows = [row for row in ws.iter_rows(min_row=2, values_only=True)]
        assert rows[0] == (0, 'Initiatief', 'Het initiatief', 'Ondertitel 1', '#3E6FA6', 'Ja', 'Ja', 'draagt bij aan', 12)
        assert rows[1][5] == 'Nee'  # isNarrow default False -> 'Nee'
        assert rows[1][6] == 'Nee'  # isProjectRole False voor de tweede kolom
        # relationLabelToNext/nodeFontSize None -> geschreven als lege string,
        # maar een leeg-tekst-cel komt bij het opnieuw inladen terug als None
        # (zelfde openpyxl-gedrag als elders in deze testsuite, geen bug).
        assert rows[1][7] is None
        assert rows[1][8] is None

    def test_kolommen_tab_gesorteerd_op_position_ongeacht_invoervolgorde(self):
        columns = make_columns(['A', 'B', 'C'])
        # Invoervolgorde expres door elkaar husselen — de tab moet toch op
        # 'position' gesorteerd zijn, niet op de volgorde van de lijst.
        shuffled = [columns[2], columns[0], columns[1]]
        wb = load(build_template_workbook('nieuw', columns=shuffled))
        types = [row[1] for row in wb['Kolommen'].iter_rows(min_row=2, values_only=True)]
        assert types == ['A', 'B', 'C']
